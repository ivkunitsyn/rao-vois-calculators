# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import io
import json
import math
import os
import sqlite3
import re
from contextlib import redirect_stdout
from decimal import Decimal, ROUND_HALF_UP
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote
import unicodedata

from functools import lru_cache
from urllib import error as urlerror
from urllib import request as urlrequest

import openpyxl
import pandas as pd


# --------------------------- helpers: progress ---------------------------

class Progress:
    def __init__(self, enabled: bool = True) -> None:
        self.enabled = enabled
        self.total = 10
        self.step = 0

    def tick(self, msg: str) -> None:
        if not self.enabled:
            return
        self.step = min(self.total, self.step + 1)
        filled = int((self.step / self.total) * 10)
        bar = "■" * filled + "□" * (10 - filled)
        pct = int((self.step / self.total) * 100)
        print(f"Прогресс: [{bar}] {pct}% — {msg}")


# --------------------------- parsing ---------------------------

def _env_flag(name: str, default: str = "0") -> bool:
    val = str(os.getenv(name, default)).strip().lower()
    return val in {"1", "true", "yes", "on"}


# По умолчанию числовое население без единиц НЕ домножается.
# Принудительный режим «числа в тысячах» можно включить через RKN_POPULATION_ASSUME_THOUSANDS=1.
RKN_POPULATION_ASSUME_THOUSANDS = _env_flag("RKN_POPULATION_ASSUME_THOUSANDS", "0")
RKN_POPULATION_ASSUME_THOUSANDS_MAX_RAW = int(os.getenv("RKN_POPULATION_ASSUME_THOUSANDS_MAX_RAW", "10000000"))
RKN_RUNTIME_POPULATION_MODE = str(os.getenv("RKN_RUNTIME_POPULATION_MODE", "yandex")).strip().lower() or "yandex"
RKN_YANDEX_API_KEY = str(os.getenv("RKN_YANDEX_API_KEY", os.getenv("YANDEX_API_KEY", ""))).strip()
# Фолбэк-ключ для приватного окружения, когда env не прокинут в runtime.
RKN_YANDEX_FALLBACK_API_KEY = str(
    os.getenv("RKN_YANDEX_FALLBACK_API_KEY", "<REDACTED>")
).strip()
RKN_YANDEX_FOLDER_ID = str(os.getenv("RKN_YANDEX_FOLDER_ID", os.getenv("YANDEX_FOLDER_ID", ""))).strip()
# Без env folder id используем рабочий folder по умолчанию (проект РКН).
RKN_YANDEX_FALLBACK_FOLDER_ID = str(os.getenv("RKN_YANDEX_FALLBACK_FOLDER_ID", "b1g1vj0i4io7gog3prlr")).strip()
RKN_YANDEX_MODEL_URI = str(os.getenv("RKN_YANDEX_MODEL_URI", "")).strip() or (
    f"gpt://{(RKN_YANDEX_FOLDER_ID or RKN_YANDEX_FALLBACK_FOLDER_ID)}/yandexgpt-lite/latest"
    if (RKN_YANDEX_FOLDER_ID or RKN_YANDEX_FALLBACK_FOLDER_ID)
    else ""
)
RKN_YANDEX_COMPLETION_URL = str(
    os.getenv("RKN_YANDEX_COMPLETION_URL", "https://llm.api.cloud.yandex.net/foundationModels/v1/completion")
).strip()
RKN_YANDEX_TIMEOUT_SECONDS = int(os.getenv("RKN_YANDEX_TIMEOUT_SECONDS", "30") or "30")


def _runtime_yandex_cfg() -> Tuple[str, str, str, int]:
    api_key = str(
        os.getenv(
            "RKN_YANDEX_API_KEY",
            os.getenv("YANDEX_API_KEY", RKN_YANDEX_API_KEY or RKN_YANDEX_FALLBACK_API_KEY),
        )
    ).strip()
    folder_id = str(
        os.getenv(
            "RKN_YANDEX_FOLDER_ID",
            os.getenv("YANDEX_FOLDER_ID", RKN_YANDEX_FOLDER_ID or RKN_YANDEX_FALLBACK_FOLDER_ID),
        )
    ).strip()
    model_uri = str(os.getenv("RKN_YANDEX_MODEL_URI", RKN_YANDEX_MODEL_URI)).strip() or (
        f"gpt://{folder_id}/yandexgpt-lite/latest" if folder_id else ""
    )
    endpoint = str(os.getenv("RKN_YANDEX_COMPLETION_URL", RKN_YANDEX_COMPLETION_URL)).strip() or RKN_YANDEX_COMPLETION_URL
    timeout = int(os.getenv("RKN_YANDEX_TIMEOUT_SECONDS", str(RKN_YANDEX_TIMEOUT_SECONDS)) or str(RKN_YANDEX_TIMEOUT_SECONDS))
    return api_key, model_uri, endpoint, timeout


def _is_active_status(raw_status: Any) -> bool:
    s = str(raw_status or "").strip().lower()
    if not s:
        return True
    negative_markers = (
        "недейств",
        "аннули",
        "прекращ",
        "приостанов",
        "истек",
    )
    if any(x in s for x in negative_markers):
        return False
    # Любые остальные непустые статусы считаем активными, чтобы не терять лицензии
    # из-за нестандартных формулировок в выгрузке РКН.
    return True

def parse_inn(raw: str) -> str:
    s = re.sub(r"\D+", "", raw or "")
    if len(s) not in (10, 12):
        raise ValueError("ИНН должен состоять из 10 или 12 цифр.")
    return s


def parse_int_like(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if math.isnan(v):
            return None
        return int(round(v))
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("\u00a0", " ")
    s = s.replace(" ", "").replace(",", ".")
    if not re.match(r"^-?\d+(\.\d+)?$", s):
        return None
    return int(round(float(s)))


def parse_population(v: Any) -> Tuple[Optional[int], List[str]]:
    notes: List[str] = []
    if v is None:
        return None, notes

    if isinstance(v, int):
        n = int(v)
        return n, notes
    if isinstance(v, float):
        if math.isnan(v):
            return None, notes
        if v < 10000 and abs(v - round(v)) > 1e-9:
            notes.append("Население было дробным числом без единиц; применена эвристика «тыс.» (×1000).")
            return int(round(v * 1000)), notes
        return int(round(v)), notes

    s = str(v).strip().lower().replace("\u00a0", " ")
    if not s:
        return None, notes

    mult = 1
    if "млн" in s:
        mult = 1_000_000
    elif "тыс" in s:
        mult = 1_000

    num = re.sub(r"[^0-9,.\- ]+", "", s)
    num = num.replace(" ", "").replace(",", ".")
    if not num:
        return None, notes

    try:
        f = float(num)
    except ValueError:
        # попытка обработать диапазоны/смешанные значения (например "10-20 тыс")
        nums = re.findall(r"\d+(?:[.,]\d+)?", s)
        if not nums:
            return None, notes
        vals = []
        for x in nums:
            try:
                vals.append(float(x.replace(",", ".")))
            except ValueError:
                continue
        if not vals:
            return None, notes
        f = max(vals)
        notes.append("Население было диапазоном/смешанным значением; взято максимальное.")

    if mult == 1 and RKN_POPULATION_ASSUME_THOUSANDS and f > 0 and f <= RKN_POPULATION_ASSUME_THOUSANDS_MAX_RAW:
        notes.append("Включён режим интерпретации численности как «тыс. человек» (×1000).")
        mult = 1_000
    elif mult == 1 and f < 10000 and abs(f - round(f)) > 1e-9:
        notes.append("Население было дробным числом без единиц; применена эвристика «тыс.» (×1000).")
        mult = 1_000

    return int(round(f * mult)), notes


def parse_population_override(v: Any) -> Tuple[Optional[int], List[str]]:
    """
    Более мягкий парсер для ручного переопределения населения пользователем.

    Практический кейс: в лицензиях РКН встречаются значения вида `146238,0`
    или `146028,325`, где дробная часть фактически означает запись в тысячах
    человек. Для ручного ввода такие значения нужно понимать как
    146 238 000 и 146 028 325 соответственно.
    """
    notes: List[str] = []
    if v is None:
        return None, notes

    if isinstance(v, str):
        s = str(v).strip().lower().replace("\u00a0", " ")
        if s:
            has_units = ("тыс" in s) or ("млн" in s)
            raw_num = re.sub(r"[^0-9,.\- ]+", "", s).replace(" ", "")
            if raw_num and re.match(r"^-?\d+[.,]\d+$", raw_num) and not has_units:
                try:
                    f = float(raw_num.replace(",", "."))
                    if f >= 1000:
                        notes.append(
                            "Ручное значение населения с десятичным разделителем интерпретировано как запись в тысячах человек (×1000)."
                        )
                        return int(round(f * 1000)), notes
                except ValueError:
                    pass

    return parse_population(v)


def _extract_yandex_text(payload: dict) -> str:
    candidates = (
        ("result", "alternatives", 0, "message", "text"),
        ("result", "alternatives", 0, "text"),
        ("alternatives", 0, "message", "text"),
        ("alternatives", 0, "text"),
    )
    for path in candidates:
        cur = payload
        ok = True
        for node in path:
            if isinstance(node, int):
                if not isinstance(cur, list) or len(cur) <= node:
                    ok = False
                    break
                cur = cur[node]
            else:
                if not isinstance(cur, dict) or node not in cur:
                    ok = False
                    break
                cur = cur[node]
        if ok and cur is not None:
            text = str(cur).strip()
            if text:
                return text
    raise ValueError("Yandex GPT не вернул текст.")


@lru_cache(maxsize=4096)
def _yandex_population_normalize(raw_value: str) -> Tuple[Optional[int], str]:
    if not raw_value:
        return None, ""
    api_key, model_uri, endpoint, timeout = _runtime_yandex_cfg()
    if not api_key or not model_uri:
        raise RuntimeError(
            "Не настроен Yandex GPT для runtime-нормализации населения "
            "(нужен RKN_YANDEX_API_KEY; modelUri берётся из RKN_YANDEX_MODEL_URI "
            "или RKN_YANDEX_FOLDER_ID)."
        )

    prompt = (
        "Преобразуй запись численности населения в людей. "
        "Если число содержит запятую или точку и нет единиц измерения, считай, что это тысячи людей. "
        "Если указаны тыс. или млн., переведи значение в людей. "
        "Если значение уже указано целым числом людей, верни его без изменений. "
        "Верни только цифры. "
        "Примеры: 146238,0 -> 146238000; 146028,325 -> 146028325; 978500 -> 978500. "
        f"Значение: {raw_value}"
    )
    body = {
        "modelUri": model_uri,
        "completionOptions": {"stream": False, "temperature": 0, "maxTokens": "16"},
        "messages": [
            {"role": "system", "text": "Ты нормализуешь запись численности населения и отвечаешь только цифрами."},
            {"role": "user", "text": prompt},
        ],
    }
    req = urlrequest.Request(
        endpoint,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Api-Key {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlrequest.urlopen(req, timeout=timeout) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urlerror.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Yandex GPT HTTP {exc.code}: {detail[:300]}") from exc

    text = _extract_yandex_text(payload)
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        raise RuntimeError(f"Yandex GPT вернул нечисловой ответ: {text!r}")
    return int(digits), text


def parse_population_runtime(v: Any, *, media_raw: Any = None, region_name: Any = None, region_text: Any = None) -> Tuple[Optional[int], List[str]]:
    raw_text = "" if v is None else str(v).strip()
    if not raw_text:
        return None, []

    if RKN_RUNTIME_POPULATION_MODE != "yandex":
        return parse_population(v)

    # Явные формы обрабатываем локально, чтобы не гонять очевидные кейсы в сеть.
    compact = raw_text.replace("\u00a0", " ").replace(" ", "")
    if re.match(r"^\d+[.,]\d+$", compact):
        f = float(compact.replace(",", "."))
        return int(round(f * 1000)), ["Численность из РКН с дробной частью интерпретирована как запись в тысячах человек (runtime)."]
    low = raw_text.lower()
    if "тыс" in low or "млн" in low:
        parsed, notes = parse_population(raw_text)
        if parsed is not None:
            notes.append("Численность из РКН с единицами измерения нормализована локально (runtime).")
            return parsed, notes

    # Спец-кейс: по федеральному охвату в РКН встречается запись в тысячах без запятой.
    region_blob = f"{region_name or ''} {region_text or ''}".lower()
    if compact.isdigit():
        val = int(compact)
        if any(x in region_blob for x in ("российск", "рф", "вся россия", "вся российская федерация")) and 100000 <= val <= 200000:
            return val * 1000, ["Численность из РКН по федеральной территории интерпретирована как запись в тысячах человек (runtime)."]
        try:
            y_val, _ = _yandex_population_normalize(raw_text)
            if y_val is not None:
                if y_val != val:
                    return y_val, [f"Численность из РКН нормализована через Yandex GPT (runtime): {raw_text} → {y_val}."]
                return y_val, []
        except Exception as exc:
            return val, [f"Yandex GPT недоступен для runtime-нормализации населения; использовано значение из РКН как есть: {exc}"]
        return val, []

    try:
        y_val, _ = _yandex_population_normalize(raw_text)
        return y_val, [f"Численность из РКН нормализована через Yandex GPT (runtime): {raw_text} → {y_val}."]
    except Exception as exc:
        parsed, notes = parse_population(raw_text)
        if parsed is not None:
            notes.append(f"Yandex GPT недоступен; применена локальная эвристика runtime: {exc}")
            return parsed, notes
        return None, [f"Не удалось нормализовать численность населения: {exc}"]


def parse_license_list(items: Optional[List[str]], only_license: Optional[str]) -> List[str]:
    out: List[str] = []
    if items:
        for raw in items:
            if raw is None:
                continue
            parts = re.split(r"[,\n]+", str(raw))
            for p in parts:
                s = p.strip()
                if s:
                    out.append(s)
    if only_license:
        s = str(only_license).strip()
        if s:
            out.append(s)

    # unique, preserve order
    seen = set()
    uniq: List[str] = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


def parse_population_by_license(items: Optional[List[str]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    if not items:
        return out
    for raw in items:
        if raw is None:
            continue
        parts = re.split(r"[\n;]+", str(raw))
        for p in parts:
            s = p.strip()
            if not s:
                continue
            if "=" not in s:
                continue
            lic, val = s.split("=", 1)
            lic = lic.strip()
            if not lic:
                continue
            pop, _ = parse_population_override(val)
            if pop is None:
                continue
            out[lic] = int(pop)
    return out


def parse_rate_by_license(items: Optional[List[str]]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    if not items:
        return out
    for raw in items:
        if raw is None:
            continue
        parts = re.split(r"[\n;]+", str(raw))
        for p in parts:
            s = p.strip()
            if not s or "=" not in s:
                continue
            lic, val = s.split("=", 1)
            lic = lic.strip()
            if not lic:
                continue
            try:
                rate = float(str(val).strip().replace(",", "."))
            except ValueError:
                continue
            if rate < 0:
                continue
            out[lic] = rate
    return out


def parse_actual_share_by_channel(items: Optional[List[str]]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    if not items:
        return out
    for raw in items:
        if raw is None:
            continue
        s = str(raw).strip()
        if not s or "=" not in s:
            continue
        key, val = s.split("=", 1)
        key = key.strip()
        if not key:
            continue
        try:
            share = float(str(val).strip().replace(",", "."))
        except ValueError:
            continue
        if 0 <= share <= 100:
            out[key] = share
    return out


def parse_internet_resources_by_license(items: Optional[List[str]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    if not items:
        return out
    for raw in items:
        if raw is None:
            continue
        parts = re.split(r"[\n;]+", str(raw))
        for s in parts:
            s = s.strip()
            if not s or "=" not in s:
                continue
            lic, val = s.split("=", 1)
            lic = lic.strip()
            if not lic:
                continue
            try:
                cnt = int(float(str(val).strip().replace(",", ".")))
            except Exception:
                continue
            if cnt < 0:
                continue
            out[lic] = cnt
    return out


def parse_internet_sites_by_license(items: Optional[List[str]]) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    if not items:
        return out
    for raw in items:
        if raw is None:
            continue
        parts = re.split(r"[\n;]+", str(raw))
        for s in parts:
            s = s.strip()
            if not s or "=" not in s:
                continue
            lic, val = s.split("=", 1)
            lic = lic.strip()
            if not lic:
                continue
            sites: List[str] = []
            for token in re.split(r"[|,]+", str(val)):
                t = token.strip()
                if t:
                    sites.append(t)
            if sites:
                out[lic] = sites
    return out


def parse_hours_week(brcst_time: Any, smi_name: Any) -> Tuple[Optional[float], List[str]]:
    notes: List[str] = []

    if brcst_time is not None and str(brcst_time).strip() != "":
        s = str(brcst_time).strip().lower()
        if "кругл" in s:
            return 168.0, notes
        n = parse_int_like(s)
        if n is not None:
            return float(n), notes
        m = re.search(r"(\d+(?:[.,]\d+)?)", s)
        if m:
            try:
                return float(m.group(1).replace(",", ".")), notes
            except Exception:
                pass

    if smi_name:
        matches = re.findall(r"\((\d{1,3}(?:[.,]\d+)?)\s*(?:ч(?:ас(?:а|ов)?)?)?\)", str(smi_name), flags=re.IGNORECASE)
        if len(matches) == 1:
            try:
                val = float(matches[0].replace(",", "."))
                if 0 < val <= 168:
                    return val, notes
            except Exception:
                pass
        elif len(matches) > 1:
            notes.append("В строке СМИ обнаружено несколько значений часов; часы из названия не использованы автоматически.")

    return None, notes


def normalize_media(sreda: Any) -> str:
    s = (str(sreda or "")).lower()
    has_air = ("эфир" in s) or ("назем" in s)
    has_cable = ("кабель" in s)
    has_univ = ("универс" in s)
    if has_univ:
        return "Одновременно в эфире и по кабелю"
    if has_air and has_cable:
        return "Одновременно в эфире и по кабелю"
    return "В эфире или по кабелю"


def clean_channel_name(name: Any) -> str:
    s = str(name or "").strip()
    if not s:
        return ""
    s = re.sub(r"\s*\(\d{1,3}(?:[.,]\d+)?\s*(?:ч(?:ас(?:а|ов)?)?)?\)\s*$", "", s, flags=re.IGNORECASE).strip()
    return s


def split_channel_tokens(name: Any) -> List[Tuple[str, Optional[float]]]:
    """
    Разбор значения СМИ/каналов из РКН.
    Поддерживает случаи вида: 'Канал A (153); Канал B (5)'.
    """
    src = str(name or "").strip()
    if not src:
        return []

    parts = [p.strip() for p in re.split(r"\s*;\s*", src) if p and p.strip()]
    if not parts:
        parts = [src]

    out: List[Tuple[str, Optional[float]]] = []
    for part in parts:
        m = re.search(r"\((\d{1,3}(?:[.,]\d+)?)\s*(?:ч(?:ас(?:а|ов)?)?)?\)\s*$", part, flags=re.IGNORECASE)
        hours: Optional[float] = None
        if m:
            try:
                val = float(m.group(1).replace(",", "."))
                if 0 < val <= 168:
                    hours = val
            except Exception:
                hours = None
        name_clean = clean_channel_name(part)
        if not name_clean:
            continue
        out.append((name_clean, hours))

    if out:
        return out
    fallback = clean_channel_name(src)
    return [(fallback, None)] if fallback else []


def is_radio_channel_name(name: Any) -> bool:
    s = str(name or "").strip().lower()
    if not s:
        return False
    s = s.replace("ё", "е")
    return bool(
        re.search(
            r"\b(радио|radio|fm|am|фм|ам|радиоканал|радиостанц)\b",
            s,
        )
    )


# --------------------------- models ---------------------------

@dataclass
class TopicShare:
    topic_raw: str
    share_pct: Optional[float]
    rate_pct: float
    note: Optional[str] = None


@dataclass
class Channel:
    name: str
    hours_week: Optional[float]
    hours_notes: List[str] = field(default_factory=list)
    topics: List[TopicShare] = field(default_factory=list)

    def avg_rate(self) -> Tuple[float, List[str]]:
        notes: List[str] = []
        if not self.topics:
            dflt = _default_topic_rate()
            notes.append(f"Тематики не найдены; ставка по умолчанию {_fmt_pct_ru(dflt)}%.")
            return dflt, notes

        shares = [t for t in self.topics if t.share_pct is not None]
        if shares:
            for t in shares:
                if t.share_pct > 50:
                    notes.append(f"Преобладающая тематика >50%: «{t.topic_raw}» ({t.share_pct}%).")
                    return t.rate_pct, notes

            total = sum(t.share_pct for t in shares)
            if total > 0:
                wavg = sum(t.share_pct * t.rate_pct for t in shares) / total
                notes.append("Ставка телеканала рассчитана как взвешенное среднее по долям тематик.")
                return wavg, notes

        avg = sum(t.rate_pct for t in self.topics) / len(self.topics)
        notes.append("Доли тематик отсутствуют/неполные; ставка телеканала рассчитана как простое среднее.")
        return avg, notes


@dataclass
class License:
    license_id: str
    org_name: str
    inn: str
    media_raw: str
    media_class: str
    population_total: Optional[int]
    population_notes: List[str] = field(default_factory=list)
    license_date: str = ""
    service_start_date: str = ""
    rate_override: Optional[float] = None
    internet_resources: int = 0
    internet_sites: List[str] = field(default_factory=list)
    rkn_url: str = ""
    channels: List[Channel] = field(default_factory=list)
    licensed_activity: str = ""

    def total_hours(self) -> float:
        hrs = [c.hours_week for c in self.channels if c.hours_week is not None]
        if hrs:
            return float(sum(hrs))
        return 168.0


# --------------------------- topic -> rate ---------------------------

DEFAULT_TOPIC_RATE = 2.5
RUNTIME_DEFAULT_TOPIC_RATE = DEFAULT_TOPIC_RATE


def _fmt_pct_ru(v: float) -> str:
    return f"{float(v):g}".replace(".", ",")


def _set_runtime_default_topic_rate(category_rate: Dict[str, float]) -> None:
    global RUNTIME_DEFAULT_TOPIC_RATE
    val = None
    if "III" in category_rate:
        val = category_rate.get("III")
    if val is None:
        for k, v in category_rate.items():
            if "неотнес" in str(k).lower():
                val = v
                break
    if val is None:
        val = DEFAULT_TOPIC_RATE
    try:
        RUNTIME_DEFAULT_TOPIC_RATE = float(val)
    except Exception:
        RUNTIME_DEFAULT_TOPIC_RATE = DEFAULT_TOPIC_RATE


def _default_topic_rate() -> float:
    return float(RUNTIME_DEFAULT_TOPIC_RATE)


TOPIC_MAP_COL_CAT = "Категория тематики использования произведений по Приложению 1"
TOPIC_MAP_COL_TOPIC = "Формулировка тематики вещания в лицензии пользователя"


def normalize_topic(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ё", "е")
    # разные тире -> дефис
    s = s.replace("—", "-").replace("–", "-")
    # unicode normalize
    s = unicodedata.normalize("NFKC", s)
    # убрать скобки/кавычки/пунктуацию в пробел
    s = re.sub(r"[\"'«»“”]", " ", s)
    s = re.sub(r"[()\[\]{}]", " ", s)
    s = re.sub(r"[^a-zа-я0-9\- ]+", " ", s)
    # схлопнуть пробелы
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Дополнительные тематики из перечня (скриншоты РАО).
EXTRA_TOPICS_BY_CATEGORY: Dict[str, List[str]] = {
    "I": [
        "аналитическое",
        "информационно-новостное",
        "вещание в сфере бизнеса",
        "военное",
        "военно-патриотическое",
        "духовно-просветительское",
        "интервью",
        "информационно-аналитическое",
        "информационно-деловое",
        "информационное",
        "информационно-публицистическое",
        "информационно-справочное",
        "информационно-экономическое",
        "информационные выпуски",
        "обзоры новостей",
        "информационные рубрики",
        "информационные передачи (программы)",
        "комментарии",
        "новостное (любой тематики)",
        "новостное",
        "общественно-информационное",
        "общественно-политическое",
        "общественно-социальное",
        "оперативная информация для водителей",
        "официальная хроника и публицистика",
        "патриотическое",
        "политическое",
        "правовое",
        "производственно-экономическое",
        "публицистическо-аналитическое",
        "публицистическое",
        "религиозное",
        "религиозно-просветительское",
        "сельскохозяйственное",
        "сообщения и объявления",
        "официальные",
        "социальное",
        "социально-значимые передачи",
        "социально-публицистическое",
        "социально-экономическое",
        "специализированное информационное",
        "справочное",
        "экологическое",
        "экономическое",
    ],
    "II": [
        "детское",
        "для детей",
        "для подростков",
        "для школьников",
        "информационно-культурное",
        "информационно-познавательное",
        "информационно-спортивное",
        "культурно-публицистическое",
        "литературно-публицистическое",
        "научное",
        "научно-образовательное",
        "научно-познавательное",
        "научно-популярное",
        "образовательное",
        "образовательные передачи (программы)",
        "подростковое",
        "познавательное",
        "пропаганда здорового образа жизни",
        "просветительское",
        "просветительско-образовательное",
        "разговорное",
        "разговорные передачи",
        "разговорные передачи или программы",
        "семейное",
        "спортивное",
        "спортивно-оздоровительное",
        "спортивные передачи (программы)",
        "туризм",
        "уроки",
        "учебно-познавательное",
        "учебно-просветительское",
        "художественно-политическое",
        "юношеское",
    ],
    "III": [
        "досуг",
        "искусство",
        "культурное",
        "культурно-просветительское",
        "культурно-развлекательное",
        "литературно-поэтическое",
        "литературно-художественное",
        "литературно-художественные программы",
        "народное творчество",
        "отдых",
        "передачи об искусстве",
        "прогноз погоды",
        "социально-культурное",
        "спортивно-развлекательное",
        "тематические передачи (программы)",
        "художественное",
        "художественно-публицистическое",
        "художественные и (или) документальные кино- и телефильмы",
        "художественные передачи (программы)",
        "документальные кино- и телефильмы",
    ],
    "IV": [
        "викторины",
        "игры",
        "информационно-развлекательное",
        "информационно-рекламное",
        "конкурсы",
        "литературно-драматическое",
        "молодежное",
        "молодежные и развлекательные передачи (программы)",
        "передачи развлекательного характера",
        "поздравительно-развлекательное",
        "поздравления",
        "познавательно-развлекательное",
        "развлекательно-воспитательное",
        "развлекательное",
        "развлекательно-игровое",
        "развлекательно-информационное",
        "развлекательно-познавательное",
        "развлекательные передачи (программы)",
        "развлечения",
        "рекламно-информационное",
        "ток-шоу",
        "ток шоу",
        "шоу",
        "шоу-программы",
        "юмористические передачи (программы)",
        "юмористическое",
    ],
    "V": [
        "концерты, в т.ч. по заявкам",
        "концерты",
        "литературно-музыкальное",
        "музыка",
        "музыкальное",
        "музыкально-информационное",
        "информационно-музыкальное",
        "музыкально-информационные передачи (программы)",
        "развлекательного характера (программы)",
        "музыкально-информационно-развлекательное",
        "музыкально-поздравительное",
        "музыкально-развлекательное",
        "музыкально-развлекательные передачи (программы)",
        "музыкально-тематические передачи",
        "музыкальные и развлекательные передачи",
        "музыкальные конкурсы",
        "музыкальные новости",
        "музыкальные передачи (программа)",
        "передачи (программы) о музыке",
        "песни",
        "популярная музыка",
        "развлекательно-музыкальное",
        "реклама",
        "рекламное",
        "рекламно-развлекательное",
        "рекламные ролики",
        "рекламные сообщения и материалы",
        "сюжеты на правах рекламы",
        "трансляция музыкальных передач (программ), концертов, фестивалей, праздников и других передач музыкального содержания",
        "трансляция музыкальных передач",
        "передачи музыкального содержания",
    ],
}


def _build_extra_topics_df() -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for cat, topics in EXTRA_TOPICS_BY_CATEGORY.items():
        for t in topics:
            t = str(t or "").strip()
            if not t:
                continue
            rows.append({
                TOPIC_MAP_COL_TOPIC: t,
                TOPIC_MAP_COL_CAT: cat,
            })
    if not rows:
        return pd.DataFrame(columns=[TOPIC_MAP_COL_TOPIC, TOPIC_MAP_COL_CAT])
    return pd.DataFrame(rows)


EXTRA_TOPICS_DF = _build_extra_topics_df()

KNOWN_TOPICS_CANON: Dict[str, str] = {}
for _cat, _topics in EXTRA_TOPICS_BY_CATEGORY.items():
    for _t in _topics:
        _norm = normalize_topic(str(_t or ""))
        if _norm and _norm not in KNOWN_TOPICS_CANON:
            KNOWN_TOPICS_CANON[_norm] = str(_t)
KNOWN_TOPICS_SORTED = sorted(KNOWN_TOPICS_CANON.keys(), key=len, reverse=True)


def _best_known_topic_match(norm_text: str) -> Optional[str]:
    if not norm_text:
        return None
    if norm_text in KNOWN_TOPICS_CANON:
        return KNOWN_TOPICS_CANON[norm_text]
    for k in KNOWN_TOPICS_SORTED:
        if len(k) < 6:
            continue
        # Только когда формулировка из перечня входит в текущий фрагмент.
        # Обратное вхождение дает шум для коротких слов типа «новости».
        if k in norm_text:
            return KNOWN_TOPICS_CANON[k]
    return None


def _canonical_topic_fragment(fragment: str, allow_unknown: bool) -> Optional[str]:
    txt = str(fragment or "").strip(" \t\r\n,;:.")
    if not txt:
        return None
    norm = normalize_topic(txt)
    if not norm:
        return None

    known = _best_known_topic_match(norm)
    if known:
        return known

    # Морфологические хвосты в длинных композитных формулировках.
    stem_rules = [
        (r"\bинформацион\w*", "информационное"),
        (r"\bобщественно[- ]полит\w*", "общественно-политическое"),
        (r"\bпросветитель\w*", "просветительское"),
        (r"\bспортив\w*", "спортивное"),
        (r"\bкультур\w*", "культурное"),
        (r"\bобразоват\w*", "образовательное"),
        (r"\bразвлек\w*", "развлекательное"),
        (r"\bмузык\w*", "музыкальное"),
        (r"\bдетск\w*", "детское"),
        (r"\bпублицист\w*", "публицистическое"),
    ]
    for pat, topic in stem_rules:
        if re.search(pat, norm):
            return topic

    if allow_unknown:
        return txt
    return None


def split_compound_topics(direction: str, decompose: bool = True) -> List[str]:
    src = str(direction or "").strip()
    if not src:
        return []
    if not decompose:
        return [src]

    is_compound = (":" in src) or (";" in src)
    if not is_compound:
        return [src]

    parts: List[str] = []
    left = src
    right = ""
    if ":" in src:
        left, right = src.split(":", 1)
        left_topic = _canonical_topic_fragment(left, allow_unknown=False)
        left_norm = normalize_topic(left_topic or left)
        # Формулировки вида «Информационное: ...» часто являются заголовком
        # перечисления подтематик; не учитываем их как отдельную тематику.
        if left_norm not in {"информационное", "информационно-новостное", "новостное"}:
            parts.append(left)
    else:
        right = src

    right = right.replace(";", ",")
    parts.extend([p for p in right.split(",") if p and p.strip()])

    out: List[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        subparts = re.split(r"\s+и\s+", p, flags=re.IGNORECASE)
        if not subparts:
            subparts = [p]
        for sub in subparts:
            topic = _canonical_topic_fragment(sub, allow_unknown=False)
            if topic:
                out.append(topic)

    # Если разумного разбиения не получилось, оставляем исходник как есть.
    if not out:
        return [src]

    uniq: List[str] = []
    seen = set()
    for t in out:
        key = normalize_topic(t)
        if not key or key in seen:
            continue
        seen.add(key)
        uniq.append(t)
    return uniq or [src]

def build_category_rate_map(vars_xlsx: Path) -> Dict[str, float]:
    df = pd.read_excel(vars_xlsx, sheet_name="Категории и ставки")
    cat_col = "Категория использования произведений (по Приложению 1)"
    rate_col = "Ставка авторского вознаграждения, процентов от дохода или расходов"
    if cat_col not in df.columns:
        for c in df.columns:
            if "Категория использования" in str(c):
                cat_col = c
                break
    if rate_col not in df.columns:
        for c in df.columns:
            if "Ставка авторского вознаграждения" in str(c):
                rate_col = c
                break
    out: Dict[str, float] = {}
    for _, r in df.iterrows():
        cat = str(r.get(cat_col, "")).strip()
        rate = r.get(rate_col)
        if cat and pd.notna(rate):
            out[cat] = float(rate)
    return out


def contract_rate_from_actual_usage_share(vars_xlsx: Path, share_pct: float) -> Tuple[Optional[float], Optional[str], str]:
    """
    Расчёт ставки договора по фактической доле использования произведений (п.2.1/п.10-12).
    Возвращает: (ставка, категория, пояснение)
    """
    try:
        df = pd.read_excel(vars_xlsx, sheet_name="Категории и ставки")
    except Exception as e:
        return None, None, f"Не удалось прочитать лист «Категории и ставки»: {e}"

    cat_col = "Категория использования произведений (по Приложению 1)"
    min_col = "Минимальная доля использования произведений, процентов"
    max_col = "Максимальная доля использования произведений, процентов"
    rate_col = "Ставка авторского вознаграждения, процентов от дохода или расходов"
    for c in df.columns:
        cs = str(c)
        if ("Категория использования" in cs) and (cat_col not in df.columns):
            cat_col = c
        if ("Минимальная доля использования" in cs) and (min_col not in df.columns):
            min_col = c
        if ("Максимальная доля использования" in cs) and (max_col not in df.columns):
            max_col = c
        if ("Ставка авторского вознаграждения" in cs) and (rate_col not in df.columns):
            rate_col = c
    required = [cat_col, min_col, max_col, rate_col]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, None, f"В листе «Категории и ставки» отсутствуют колонки: {missing}"

    x = float(share_pct)
    for _, r in df.iterrows():
        try:
            lo = float(r[min_col])
            hi = float(r[max_col])
            rate = float(r[rate_col])
            cat = str(r[cat_col]).strip()
        except Exception:
            continue
        if not math.isfinite(rate):
            continue
        if not math.isfinite(lo):
            continue
        if not math.isfinite(hi):
            hi = float("inf")
        if lo <= x <= hi:
            hi_txt = "и выше" if math.isinf(hi) else f"{hi:g}%"
            return round_rate(rate), cat, f"Фактическая доля {x:g}% попадает в диапазон {lo:g}%–{hi_txt} (категория {cat})."

    # Фолбэк: берём ближайшую верхнюю/нижнюю границу.
    rows = []
    for _, r in df.iterrows():
        try:
            lo = float(r[min_col])
            hi = float(r[max_col])
            rate = float(r[rate_col])
            cat = str(r[cat_col]).strip()
            if not math.isfinite(rate):
                continue
            if not math.isfinite(lo):
                continue
            if not math.isfinite(hi):
                hi = float("inf")
            rows.append((lo, hi, rate, cat))
        except Exception:
            continue
    if not rows:
        return None, None, "В листе «Категории и ставки» нет валидных строк."
    rows = sorted(rows, key=lambda t: t[0])
    if x < rows[0][0]:
        lo, hi, rate, cat = rows[0]
        hi_txt = "и выше" if math.isinf(hi) else f"{hi:g}%"
        return round_rate(rate), cat, f"Фактическая доля {x:g}% ниже таблицы; применён минимальный диапазон {lo:g}%–{hi_txt} (категория {cat})."
    lo, hi, rate, cat = rows[-1]
    hi_txt = "и выше" if math.isinf(hi) else f"{hi:g}%"
    return round_rate(rate), cat, f"Фактическая доля {x:g}% выше таблицы; применён максимальный диапазон {lo:g}%–{hi_txt} (категория {cat})."


def topic_to_rate(
    topic: str,
    category_rate: Dict[str, float],
    mapping_df: Optional[pd.DataFrame],
    *,
    strict_topic_match: bool = False,
) -> Tuple[float, List[str]]:
    notes: List[str] = []

    t_raw = (topic or "").strip()
    tl = normalize_topic(t_raw)

    # Формулировки вида «Культурно-просветительское (армянская и зарубежная музыка)»
    # считаем уточнением известной базовой тематики, а не составной тематикой.
    head_topic = re.split(r"\s*[\(\[]", t_raw, maxsplit=1)[0].strip(" \t\r\n,;:.")
    head_norm = normalize_topic(head_topic)
    if head_topic and head_topic != t_raw and head_norm in KNOWN_TOPICS_CANON:
        canonical_head = KNOWN_TOPICS_CANON[head_norm]
        if normalize_topic(canonical_head) != tl:
            rate, head_notes = topic_to_rate(
                canonical_head,
                category_rate,
                mapping_df,
                strict_topic_match=strict_topic_match,
            )
            return rate, [
                f"Уточнение в скобках не меняет базовую тематику; применена ставка для «{canonical_head}».",
                *head_notes,
            ]

    # Комбинированная тематика одной строкой без долей (через запятые/союзы):
    # применяем максимальную ставку из распознанных категорий.
    # Это закрывает кейсы вида:
    # «Информационно-новостная ... культурной, спортивной, образовательной, политической, музыкальной ...».
    if tl and ("," in t_raw or " и " in tl):
        cats: List[str] = []
        if re.search(r"\b(информац|новост|аналит|публицист|полит|эконом|религи|прав)\w*", tl):
            cats.append("I")
        if re.search(r"\b(дет|образоват|науч|познав|спорт|оздоров|зож|туризм)\w*", tl):
            cats.append("II")
        if re.search(r"\b(культур|искусств|литератур|художествен|документ|кино|просвет)\w*", tl):
            cats.append("III")
        if re.search(r"\b(развлек|юмор|шоу|викторин|конкурс|молодеж|игров)\w*", tl):
            cats.append("IV")
        if re.search(r"\b(музык|песн|клип|концерт|эстрад|реклам)\w*", tl):
            cats.append("V")
        cats = list(dict.fromkeys(cats))
        if len(cats) >= 2:
            best_cat = max(cats, key=lambda c: float(category_rate.get(c, _default_topic_rate())))
            best_rate = float(category_rate.get(best_cat, _default_topic_rate()))
            notes.append(
                f"Комбинированная тематика без долей: применена максимальная ставка из распознанных категорий ({best_cat})."
            )
            return best_rate, notes

    # ---------------- 1) Пытаемся сопоставить по таблице (не только ==, но и contains) ----------------
    if mapping_df is not None and not mapping_df.empty:
        df = mapping_df.copy()
    else:
        df = pd.DataFrame(columns=[TOPIC_MAP_COL_TOPIC, TOPIC_MAP_COL_CAT])

    if not EXTRA_TOPICS_DF.empty:
        df = pd.concat([df, EXTRA_TOPICS_DF], ignore_index=True)

    if TOPIC_MAP_COL_CAT in df.columns and TOPIC_MAP_COL_TOPIC in df.columns:
        df = df[[TOPIC_MAP_COL_TOPIC, TOPIC_MAP_COL_CAT]].dropna()
        if not df.empty:
            df = df.copy()
            df["_norm"] = df[TOPIC_MAP_COL_TOPIC].astype(str).map(normalize_topic)

            # 1) точное совпадение
            m = df[df["_norm"] == tl]
            if not m.empty:
                cat = str(m.iloc[0][TOPIC_MAP_COL_CAT]).strip()
                rate = category_rate.get(cat)
                if rate is not None:
                    notes.append(f"Тематика сопоставлена по таблице «Тематики по категориям» (точно): категория {cat}.")
                    return float(rate), notes

            if strict_topic_match:
                dflt = _default_topic_rate()
                notes.append(
                    f"Точная формулировка тематики не найдена в таблице; применена средняя ставка {_fmt_pct_ru(dflt)}%."
                )
                return dflt, notes

            # 2) “вхождение” в обе стороны (часто формулировки длиннее/короче)
            # берем самый “длинный” матч (обычно он точнее)
            # 2) “вхождение” в обе стороны (часто формулировки длиннее/короче)
            # 2) “вхождение” в обе стороны (длиннее/короче)
            # mask1: в лицензии tl содержит формулировку из таблицы
            # mask2: формулировка из таблицы содержит tl
            if tl:  # защита от пустой строки
                # сначала ищем формулировки, которые входят в тематику (точнее)
                mask_in_tl = df["_norm"].apply(lambda x: bool(x) and (x in tl))
                candidates = df[mask_in_tl].copy()
                if candidates.empty:
                    # затем ищем тематики, которые входят в формулировку (менее точное совпадение)
                    mask_tl_in = df["_norm"].str.contains(tl, na=False, regex=False)
                    candidates = df[mask_tl_in].copy()

                if not candidates.empty:
                    candidates["_len"] = candidates["_norm"].str.len()
                    candidates = candidates.sort_values("_len", ascending=False)

                    cat = str(candidates.iloc[0][TOPIC_MAP_COL_CAT]).strip()
                    rate = category_rate.get(cat)
                    if rate is not None:
                        notes.append(f"Тематика сопоставлена по таблице «Тематики по категориям» (вхождение): категория {cat}.")
                        return float(rate), notes



    if strict_topic_match:
        dflt = _default_topic_rate()
        notes.append(
            f"Точная формулировка тематики не найдена в таблице; применена средняя ставка {_fmt_pct_ru(dflt)}%."
        )
        return dflt, notes

    # ---------------- 2) Эвристики (расширенные) ----------------
    def hit(*keys: str) -> bool:
        return any(k in tl for k in keys)

    # Составные формулировки без долей (например, «познавательно-развлекательные...»):
    # если внутри одной тематики явно смешаны несколько категорий, берём ставку
    # по наиболее «тяжёлой» категории (максимальная ставка).
    connector_like = ("-" in tl) or (" и " in tl) or ("/" in tl)
    matched_categories: List[str] = []
    if hit("информационно-", "информацион", "новост", "аналит", "публицист", "полит", "эконом"):
        matched_categories.append("I")
    if hit("дет", "образоват", "учеб", "урок", "научн", "познавательн", "спорт", "оздоров", "здоров", "зож", "туризм"):
        matched_categories.append("II")
    if hit("культур", "искусств", "литератур", "поэтич", "художественн", "документ", "кино", "телефильм", "погода", "отдых", "досуг"):
        matched_categories.append("III")
    if hit("развлек", "юмор", "шоу", "ток-шоу", "ток шоу", "викторин", "конкурс", "поздрав", "комед", "розыгрыш", "молодеж", "игры", "игровое"):
        matched_categories.append("IV")
    if hit("музык", "песн", "клип", "концерт", "эстрад", "музыка"):
        matched_categories.append("V")

    unique_categories = list(dict.fromkeys(matched_categories))
    if connector_like and len(unique_categories) >= 2:
        best_cat = max(unique_categories, key=lambda c: float(category_rate.get(c, _default_topic_rate())))
        best_rate = float(category_rate.get(best_cat, _default_topic_rate()))
        notes.append(
            f"Составная тематика с несколькими категориями; применена максимальная ставка по категории {best_cat}."
        )
        return best_rate, notes

    # ВАЖНО: порядок правил — это качество классификации.
    # Сначала “составные” и более специфичные штуки, потом общие.

    # Если тематика формата «... информация» не сопоставилась с Приложением №1,
    # применяем дефолтную ставку (категория по умолчанию), чтобы не уводить в I по общим подстрокам.
    if re.search(r"\bинформац(?:ия|ии|ию|ией)\b", tl) and "информационно-" not in tl:
        dflt = _default_topic_rate()
        notes.append(f"Тематика не сопоставлена с Приложением №1 (формат «... информация»); применена ставка {_fmt_pct_ru(dflt)}%.")
        return dflt, notes

    # ---- V (музыкальная) — ловим раньше, чтобы “информационно-музыкальное” не ушло в I
    if hit("музык", "песн", "клип", "концерт", "эстрад", "популярная музыка", "музыкально", "музыка"):
        rate = category_rate.get("V", 3.0)
        notes.append("Тематика распознана эвристикой как «музыкальная» (категория V).")
        return float(rate), notes

    # ---- II (социально-полезная/образование/дети/спорт/наука/здоровье/туризм/уроки)
    if hit(
        "дет", "для детей", "подрост", "школьник",
        "образоват", "учеб", "урок", "просветительско-образователь", "научн", "научно",
        "познавательн", "учебно-познавательн", "учебно-просветительск",
        "спорт", "спортивно", "оздоров", "здоров", "зож", "пропаганда здорового",
        "туризм"
    ):
        rate = category_rate.get("II", 2.3)
        notes.append("Тематика распознана эвристикой как «социально-полезная/образовательная/спорт/ЗОЖ» (категория II).")
        return float(rate), notes

    # ---- III (культурно-просветительская/искусство/документалистика/художественное/погода/отдых/народное)
    if hit(
        "культур", "искусств", "литератур", "поэтич", "художественн", "художественно",
        "документ", "кино", "телефильм", "передачи об искусстве",
        "народное творчество", "социально-культурн",
        "прогноз погоды", "погода", "отдых", "досуг"
    ):
        rate = category_rate.get("III", _default_topic_rate())
        notes.append("Тематика распознана эвристикой как «культурно-просветительская/художественная» (категория III).")
        return float(rate), notes

    # ---- IV (развлекательная) — проверяем после III, чтобы не перехватывать
    # художественные/игровые кинофильмы (категория III).
    if hit(
        "развлек", "юмор", "шоу", "ток-шоу", "ток шоу",
        "викторин", "конкурс", "поздрав", "комед", "коморист", "розыгрыш", "молодеж",
        "игры", "игровое", "развлекательно-игровое"
    ):
        rate = category_rate.get("IV", 2.7)
        notes.append("Тематика распознана эвристикой как «развлекательная» (категория IV).")
        return float(rate), notes

    # ---- I (информационная/новости/политика/экономика/право/религия/официальная хроника/социально-значимые)
    if hit(
        "информационно-", "информационное", "новост", "аналит", "публицист", "общественно-",
        "делов", "правов", "социально-значим", "официальная хроника",
        "оперативная информация для водителей", "религи", "патриот", "интервью", "комментар",
        "политическое", "экономическое"
    ):
        rate = category_rate.get("I", 2.0)
        notes.append("Тематика распознана эвристикой как «информационная» (категория I).")
        return float(rate), notes

    # Если вообще непонятно — дефолт
    dflt = _default_topic_rate()
    notes.append(f"Тематика не распознана; применена ставка по умолчанию {_fmt_pct_ru(dflt)}% (категория III).")
    return dflt, notes


# --------------------------- loading: RKN table ---------------------------

def iter_rkn_rows(rkn_xlsx: Path) -> Tuple[List[str], Any]:
    wb = openpyxl.load_workbook(rkn_xlsx, read_only=True, data_only=True)
    ws = wb.active

    header_raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    while header_raw and (header_raw[-1] is None or str(header_raw[-1]).strip() == ""):
        header_raw.pop()
    header = header_raw
    max_col = len(header)

    it = ws.iter_rows(min_row=2, max_col=max_col, values_only=True)
    return header, it


def _resolve_rkn_sqlite_for_xlsx(rkn_xlsx: Path) -> Optional[Path]:
    env_path = str(os.getenv("RKN_DB_PATH", "") or "").strip()
    if env_path:
        p = Path(env_path).expanduser()
        if p.exists():
            return p
        return None
    candidates = [
        rkn_xlsx.with_suffix(".sqlite"),
        rkn_xlsx.with_suffix(".db"),
        rkn_xlsx.parent / "Таблица РКН.sqlite",
        rkn_xlsx.parent / "Таблица РКН.db",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _matching_rkn_rows_from_sqlite(db_path: Path, inn: str) -> Optional[Tuple[Tuple[Tuple[Any, ...], ...], Tuple[str, ...]]]:
    conn = sqlite3.connect(str(db_path))
    try:
        cols = {str(r[1]): int(r[0]) for r in conn.execute("PRAGMA table_info(rkn)")}
        required_cols = {
            "inn", "org_name", "license_num", "sreda", "population",
            "region_name_full", "region_text", "smi_name14", "smi_name",
            "brcst_direction", "percentage", "brcst_time", "status",
        }
        if not required_cols.issubset(cols):
            return None
        has_activity = "licensed_activity" in cols
        has_license_date = "license_date" in cols
        has_service_start_date = "service_start_date" in cols
        has_freq = "freq" in cols
        has_brcst_descr = "brcst_descr" in cols
        rows = conn.execute(
            """
            SELECT org_name, license_num, sreda, population, region_name_full, region_text,
                   smi_name14, smi_name, brcst_time, brcst_direction, percentage,
                   """
            + ("license_date, " if has_license_date else "'' AS license_date, ")
            + ("service_start_date, " if has_service_start_date else "'' AS service_start_date, ")
            + ("freq, " if has_freq else "'' AS freq, ")
            + ("brcst_descr, " if has_brcst_descr else "'' AS brcst_descr, ")
            + """
                   """
            + ("licensed_activity, " if has_activity else "'' AS licensed_activity, ")
            + """
                   status
            FROM rkn WHERE inn = ?
            """,
            (str(inn).strip(),),
        ).fetchall()
    finally:
        conn.close()

    out: List[Tuple[Any, ...]] = []
    for row in rows:
        status = row[16]
        if not _is_active_status(status):
            continue
        out.append(row[:16])
    return tuple(out), tuple()

@lru_cache(maxsize=1)
def _inn_to_org_map(rkn_path: str, mtime: float) -> dict:
    rkn_xlsx = Path(rkn_path)
    header, it = iter_rkn_rows(rkn_xlsx)
    idx = {h: i for i, h in enumerate(header)}

    col_inn = idx.get("ns1:inn")
    col_name = idx.get("ns1:org_name")
    if col_inn is None or col_name is None:
        return {}

    out = {}
    for row in it:
        inn = str(row[col_inn] or "").strip()
        if not inn:
            continue
        if inn not in out:
            out[inn] = str(row[col_name] or "").strip()
    return out

def get_org_name_by_inn(rkn_xlsx: Path, inn: str) -> str:
    mp = _inn_to_org_map(str(rkn_xlsx), rkn_xlsx.stat().st_mtime)
    return (mp.get(inn) or "").strip()



def build_rkn_url(license_id: str) -> str:
    return "https://rkn.gov.ru/activity/mass-media/for-broadcasters/teleradio/?id=" + quote(str(license_id), safe="")


def _has_freq_value(raw: Any) -> bool:
    return bool(str(raw or "").strip())


def _aggregate_radio_population_from_rows(
    pop_rows: List[Dict[str, Any]],
    *,
    runtime_population_normalization: bool,
    media_raw: Any = None,
) -> Tuple[Optional[int], List[str]]:
    if not pop_rows:
        return None, []

    rows_with_freq = [row for row in pop_rows if _has_freq_value(row.get("freq"))]
    if not rows_with_freq:
        return None, []

    total = 0
    notes: List[str] = []
    for row in rows_with_freq:
        raw = row.get("population")
        if runtime_population_normalization:
            pop_int, pop_notes = parse_population_runtime(
                raw,
                media_raw=row.get("brcst_descr") or media_raw,
                region_name=row.get("region_name_full"),
                region_text=row.get("region_text"),
            )
        else:
            pop_int, pop_notes = parse_population(raw)
        if pop_int is not None:
            total += int(pop_int)
        if pop_notes:
            notes.extend(pop_notes)
    if total <= 0:
        return None, notes[:2]
    return int(total), notes[:2]


def load_licenses_by_inn(
    rkn_xlsx: Path,
    inn: str,
    vars_xlsx: Path,
    *,
    filter_radio_channels: bool = True,
    runtime_population_normalization: bool = True,
    strict_topic_match: bool = False,
) -> Tuple[List[License], List[str]]:
    notes: List[str] = []

    xlsx_mtime = float(rkn_xlsx.stat().st_mtime)
    vars_mtime = float(vars_xlsx.stat().st_mtime)
    rows, missing = _matching_rkn_rows_cached(str(rkn_xlsx), xlsx_mtime, inn)
    if missing:
        notes.append(f"В таблице РКН не найдены ожидаемые колонки: {list(missing)}. Скрипт будет работать частично.")

    category_rate = _category_rate_map_cached(str(vars_xlsx), vars_mtime)
    _set_runtime_default_topic_rate(category_rate)
    topics_map = _topics_map_cached(str(vars_xlsx), vars_mtime)

    by_license: Dict[str, Dict[str, Any]] = {}

    for row in rows:
        org_name = str(row[0] or "").strip()
        lic_id = str(row[1] or "").strip()
        sreda = str(row[2] or "").strip()
        pop_raw = row[3]
        region_name_full = str(row[4] or "").strip()
        region_text = str(row[5] or "").strip()
        smi14_raw = row[6]
        smi_raw = row[7]
        channel_tokens = split_channel_tokens(smi14_raw) or split_channel_tokens(smi_raw)
        if not channel_tokens:
            smi14 = clean_channel_name(smi14_raw)
            smi = clean_channel_name(smi_raw)
            channel_tokens = [(smi14 or smi or "Неизвестный канал", None)]
        brcst_time = row[8]
        direction = str(row[9] or "").strip()
        perc = row[10]
        license_date = str((row[11] if len(row) > 11 else "") or "").strip()
        service_start_date = str((row[12] if len(row) > 12 else "") or "").strip()
        freq = row[13] if len(row) > 13 else None
        brcst_descr = str((row[14] if len(row) > 14 else "") or "").strip()
        licensed_activity = str((row[15] if len(row) > 15 else "") or "").strip()

        if not lic_id:
            continue

        lic = by_license.setdefault(lic_id, {
            "org_name": org_name,
            "inn": inn,
            "sreda": sreda,
            "licensed_activity": licensed_activity,
            "license_date": license_date,
            "service_start_date": service_start_date,
            "region_name_full": region_name_full,
            "region_text": region_text,
            "pop_values": [],
            "pop_notes": [],
            "pop_rows": [],
            "channels": {}
        })
        if not lic.get("licensed_activity") and licensed_activity:
            lic["licensed_activity"] = licensed_activity
        if not lic.get("license_date") and license_date:
            lic["license_date"] = license_date
        if not lic.get("service_start_date") and service_start_date:
            lic["service_start_date"] = service_start_date

        if runtime_population_normalization:
            pop_int, pop_notes = parse_population_runtime(
                pop_raw,
                media_raw=sreda,
                region_name=region_name_full or lic.get("region_name_full"),
                region_text=region_text or lic.get("region_text"),
            )
        else:
            pop_int, pop_notes = parse_population(pop_raw)
        if pop_int is not None:
            lic["pop_values"].append(pop_int)
        lic["pop_notes"].extend(pop_notes)
        lic["pop_rows"].append({
            "population": pop_raw,
            "freq": freq,
            "brcst_descr": brcst_descr or sreda,
            "region_name_full": region_name_full,
            "region_text": region_text,
        })

        hrs, hrs_notes = parse_hours_week(brcst_time, row[7])

        if direction and len(channel_tokens) > 1:
            notes.append(
                f"Лицензия {lic_id}: строка с несколькими СМИ и тематикой «{direction}». Тематика применена ко всем указанным СМИ, проверьте вручную."
            )

        for channel_name, token_hours in channel_tokens:
            ch = lic["channels"].setdefault(channel_name, {
                "hours": None,
                "hours_notes": [],
                "topics": [],
                "is_radio": is_radio_channel_name(channel_name),
            })

            candidate_hours = token_hours if token_hours is not None else hrs
            if candidate_hours is not None:
                if ch["hours"] is None:
                    ch["hours"] = float(candidate_hours)
                else:
                    old_h = float(ch["hours"])
                    new_h = float(candidate_hours)
                    if abs(old_h - new_h) > 1e-6:
                        # При конфликтах сохраняем наибольший объём и фиксируем примечание.
                        ch["hours"] = max(old_h, new_h)
                        conflict_note = (
                            f"Обнаружены разные значения часов вещания ({old_h:g} и {new_h:g}); использовано {ch['hours']:g}."
                        )
                        if conflict_note not in ch["hours_notes"]:
                            ch["hours_notes"].append(conflict_note)
            if hrs_notes:
                for n in hrs_notes:
                    if n not in ch["hours_notes"]:
                        ch["hours_notes"].append(n)

            if direction:
                share = None
                if perc is not None and str(perc).strip() != "":
                    try:
                        share = float(str(perc).replace(",", "."))
                    except ValueError:
                        share = None

                topics = split_compound_topics(direction, decompose=not strict_topic_match)
                topic_share = share if len(topics) == 1 else None
                if len(topics) > 1 and share is not None:
                    notes.append(
                        f"Лицензия {lic_id}: в составной тематике «{direction}» доля {share}% не распределялась по под-тематикам."
                    )
                for topic_raw in topics:
                    rate, rate_notes = topic_to_rate(
                        topic_raw,
                        category_rate,
                        topics_map,
                        strict_topic_match=strict_topic_match,
                    )
                    note = "; ".join(rate_notes) if rate_notes else None
                    ch["topics"].append(TopicShare(topic_raw=topic_raw, share_pct=topic_share, rate_pct=rate, note=note))

    licenses: List[License] = []
    for lic_id, data in by_license.items():
        media_class = normalize_media(data.get("sreda"))
        channel_items = list(data["channels"].items())
        has_non_radio_channel = any(not bool(ch_data.get("is_radio")) for _, ch_data in channel_items)
        if not channel_items:
            has_non_radio_channel = True
        activity_is_radio = "радио" in str(data.get("licensed_activity") or "").lower()
        if not activity_is_radio:
            media_l = str(data.get("sreda") or "").lower()
            activity_is_radio = ("радио" in media_l) or ("радиоканал" in media_l) or ("радиовещ" in media_l)
        if not activity_is_radio and channel_items:
            activity_is_radio = all(bool(ch_data.get("is_radio")) for _, ch_data in channel_items)
        radio_pop_total, radio_pop_notes = _aggregate_radio_population_from_rows(
            list(data.get("pop_rows") or []),
            runtime_population_normalization=runtime_population_normalization,
            media_raw=data.get("sreda"),
        )
        if activity_is_radio and radio_pop_total is not None:
            pop_total = radio_pop_total
            data["pop_notes"] = radio_pop_notes
        else:
            pop_total = None
            if data["pop_values"]:
                pop_candidates = sorted(set(data["pop_values"]))
                pop_total = int(max(pop_candidates))
                if len(pop_candidates) > 1:
                    data.setdefault("pop_notes", []).append(
                        "Обнаружено несколько значений населения по лицензии; для расчёта использовано максимальное значение."
                    )
        if filter_radio_channels and not has_non_radio_channel:
            notes.append(f"Лицензия {lic_id} исключена из ТВ-расчёта: в строках РКН определён только радиоканал.")
            continue

        lic_obj = License(
            license_id=lic_id,
            org_name=data.get("org_name", ""),
            inn=inn,
            media_raw=data.get("sreda", ""),
            media_class=media_class,
            population_total=pop_total,
            population_notes=data.get("pop_notes", []),
            license_date=str(data.get("license_date") or ""),
            service_start_date=str(data.get("service_start_date") or ""),
            rkn_url=build_rkn_url(lic_id),
            channels=[],
            licensed_activity=str(data.get("licensed_activity") or ""),
        )
        for ch_name, ch_data in channel_items:
            if filter_radio_channels and has_non_radio_channel and bool(ch_data.get("is_radio")):
                continue
            lic_obj.channels.append(Channel(
                name=ch_name,
                hours_week=ch_data.get("hours"),
                hours_notes=ch_data.get("hours_notes", []),
                topics=ch_data.get("topics", [])
            ))
        licenses.append(lic_obj)

    if not licenses:
        notes.append("По этому ИНН в таблице РКН не найдено строк. Проверьте, что ИНН есть в выгрузке.")
    return licenses, notes


@lru_cache(maxsize=8)
def _category_rate_map_cached(vars_xlsx_path: str, vars_mtime: float) -> Dict[str, float]:
    return build_category_rate_map(Path(vars_xlsx_path))


@lru_cache(maxsize=8)
def _topics_map_cached(vars_xlsx_path: str, vars_mtime: float) -> pd.DataFrame:
    vars_xlsx = Path(vars_xlsx_path)
    try:
        topics_map = pd.read_excel(vars_xlsx, sheet_name="Тематики по категориям")
        if topics_map.dropna(how="all").empty:
            return pd.DataFrame()
        return topics_map
    except Exception:
        return pd.DataFrame()


@lru_cache(maxsize=512)
def _matching_rkn_rows_cached(rkn_xlsx_path: str, rkn_mtime: float, inn: str) -> Tuple[Tuple[Tuple[Any, ...], ...], Tuple[str, ...]]:
    rkn_xlsx = Path(rkn_xlsx_path)
    db_path = _resolve_rkn_sqlite_for_xlsx(rkn_xlsx)
    if db_path is not None:
        try:
            db_rows = _matching_rkn_rows_from_sqlite(db_path, inn)
            if db_rows is not None:
                return db_rows
        except Exception:
            pass
    header, it = iter_rkn_rows(rkn_xlsx)
    idx = {h: i for i, h in enumerate(header)}

    required = [
        "ns1:inn", "ns1:org_name", "ns1:license_num", "ns1:sreda", "ns1:population",
        "ns1:region_name_full", "ns1:region_text",
        "ns1:smi_name14", "ns1:smi_name", "ns1:brcst_direction", "ns1:percentage", "ns1:brcst_time",
        "ns1:status"
    ]
    missing = tuple(c for c in required if c not in idx)

    def get(row, col):
        j = idx.get(col)
        if j is None or j >= len(row):
            return None
        return row[j]

    out: List[Tuple[Any, ...]] = []
    for row in it:
        row_inn = str(get(row, "ns1:inn") or "").strip()
        if row_inn != inn:
            continue
        status = get(row, "ns1:status")
        if not _is_active_status(status):
            continue
        out.append((
            get(row, "ns1:org_name"),
            get(row, "ns1:license_num"),
            get(row, "ns1:sreda"),
            get(row, "ns1:population"),
            get(row, "ns1:region_name_full"),
            get(row, "ns1:region_text"),
            get(row, "ns1:smi_name14"),
            get(row, "ns1:smi_name"),
            get(row, "ns1:brcst_time"),
            get(row, "ns1:brcst_direction"),
            get(row, "ns1:percentage"),
            get(row, "ns1:license_date"),
            get(row, "ns1:service_start_date"),
            get(row, "ns1:freq"),
            get(row, "ns1:brcst_descr"),
            get(row, "ns1:licensed_activity"),
        ))
    return tuple(out), missing


def warm_license_cache_for_inn(rkn_xlsx: Path, inn: str) -> None:
    try:
        _matching_rkn_rows_cached(str(rkn_xlsx), float(Path(rkn_xlsx).stat().st_mtime), str(inn).strip())
    except Exception:
        return


# --------------------------- computations ---------------------------

def round_rate(x: float) -> float:
    return round(x + 1e-9, 1)


def round_rub(x: float) -> float:
    # Округление до рубля: 0-49 коп. вниз, 50-99 коп. вверх.
    return float(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def compute_license_rate(lic: License) -> Tuple[float, Dict[str, Any]]:
    if lic.rate_override is not None:
        return round_rate(float(lic.rate_override)), {
            "manual_override": True,
            "license_rate": round_rate(float(lic.rate_override)),
            "note": "Ставка по лицензии задана вручную пользователем."
        }

    det: Dict[str, Any] = {"channels": []}
    num = 0.0
    den = 0.0

    for ch in lic.channels:
        ch_rate, ch_notes = ch.avg_rate()
        hrs = ch.hours_week if ch.hours_week is not None else 168.0

        det["channels"].append({
            "channel": ch.name,
            "hours": hrs,
            "hours_notes": list(ch.hours_notes or []),
            "channel_rate_raw": ch_rate,
            "channel_rate": round_rate(ch_rate),
            "notes": ch_notes,
            "topics": [
                {"topic": t.topic_raw, "share_pct": t.share_pct, "rate_pct": t.rate_pct, "note": t.note}
                for t in ch.topics
            ]
        })

        num += ch_rate * hrs
        den += hrs

    if den == 0:
        dflt = _default_topic_rate()
        return dflt, {"warning": f"Не удалось рассчитать ставку по ВЛ (нет часов/каналов). Применена {_fmt_pct_ru(dflt)}%."}
    return round_rate(num / den), det


def compute_contract_rate(licenses: List[License]) -> Tuple[float, Dict[str, Any]]:
    details: Dict[str, Any] = {"licenses": []}
    num = 0.0
    den = 0.0

    for lic in licenses:
        lic_rate, lic_rate_details = compute_license_rate(lic)
        pop = lic.population_total
        hrs = lic.total_hours()
        w = (pop or 0) * hrs

        details["licenses"].append({
            "license_id": lic.license_id,
            "license_rate": lic_rate,
            "population": pop,
            "hours": hrs,
            "weight": w,
            "license_rate_details": lic_rate_details
        })

        if pop is None:
            continue
        num += lic_rate * w
        den += w

    if den == 0:
        dflt = _default_topic_rate()
        return dflt, {"warning": f"Не удалось рассчитать взвешенную ставку (нет населения). Применена {_fmt_pct_ru(dflt)}%."}
    return round_rate(num / den), details


def apply_actual_share_by_channel(
    licenses: List[License],
    vars_xlsx: Path,
    actual_share_by_channel: Dict[str, float],
    usage_obj_gen: str,
) -> List[str]:
    notes: List[str] = []
    for lic in licenses:
        for i, ch in enumerate(lic.channels):
            key_by_name = f"{lic.license_id}|{ch.name}"
            key_by_index = f"{lic.license_id}|{i + 1}"
            if key_by_name in actual_share_by_channel:
                share = float(actual_share_by_channel[key_by_name])
            elif key_by_index in actual_share_by_channel:
                share = float(actual_share_by_channel[key_by_index])
            else:
                continue
            factual_rate, factual_cat, factual_note = contract_rate_from_actual_usage_share(vars_xlsx, share)
            if factual_rate is None:
                notes.append(
                    f"Канал/СМИ {ch.name} ({lic.license_id}): не удалось применить фактическую долю {share:g}%: {factual_note}"
                )
                continue
            ch.topics = [
                TopicShare(
                    topic_raw=f"Фактическая доля использования {usage_obj_gen}: {share:g}%",
                    share_pct=share,
                    rate_pct=float(factual_rate),
                    note=f"Категория {factual_cat or '—'}; {factual_note}",
                )
            ]
            notes.append(
                f"Канал/СМИ {ch.name} ({lic.license_id}): ставка {float(factual_rate):.1f}% рассчитана по фактической доле {share:g}%."
            )
    return notes


def build_minimum_breakdown_lines(min_details: Dict[str, Any]) -> List[str]:
    lines: List[str] = []
    prev_min_after: Optional[float] = None
    for st in (min_details.get("steps") or []):
        step = str(st.get("step") or "").upper()
        if step == "C1+C2(+C7)":
            parts: List[str] = []
            for row in (st.get("per_license") or []):
                lic_id = str(row.get("license_id") or "—")
                pop = row.get("population")
                media = str(row.get("media") or "—")
                min_table = float(row.get("min_table") or 0.0)
                coeff = float(row.get("hour_coeff") or 1.0)
                min_after = float(row.get("min_after") or 0.0)
                coeff_part = f" × {str(coeff).replace('.', ',')}" if abs(coeff - 1.0) > 1e-9 else ""
                lines.append(
                    f"• {lic_id}: население {money(pop)}, среда «{media}» → {money(min_table)} ₽{coeff_part} = {money(min_after)} ₽."
                )
                parts.append(money(min_after))
            if parts:
                lines.append(f"Итого минимальная сумма: {' + '.join(parts)} = {money(st.get('min_after'))} ₽.")
        elif step == "C3":
            k_hours = float(st.get("hour_coeff") or 1.0)
            hour_part = f" × {str(k_hours).replace('.', ',')}" if abs(k_hours - 1.0) > 1e-9 else ""
            lines.append(
                f"• Ветка малого дохода: суммарное население {money(st.get('N_sum'))}, "
                f"среда «{st.get('media') or '—'}» → {money(st.get('min_table'))} ₽{hour_part} × {str(st.get('k_small')).replace('.', ',')} = {money(st.get('min_after'))} ₽."
            )
        elif step == "C4":
            coeff = float(st.get("coeff") or 1.0)
            if prev_min_after is not None and abs(coeff - 1.0) > 1e-9:
                lines.append(
                    f"• Скидка по количеству лицензий: {st.get('n_licenses')} лицензий → {money(prev_min_after)} ₽ × {str(coeff).replace('.', ',')} = {money(st.get('min_after'))} ₽."
                )
        elif step == "GUILLOTINE":
            if bool(st.get("triggered")):
                threshold_coef = float(st.get("threshold_coef") or 1.1)
                lines.append(
                    f"• Проверка по п.3.7: {money(st.get('min_before'))} ₽ > {str(round(threshold_coef, 2)).replace('.', ',')} × {money_precise(st.get('s_percent'))} ₽ = {money(st.get('threshold'))} ₽."
                )
                if st.get("alt_sum_population") is not None:
                    lines.append(
                        f"• Минимальная сумма по суммарной численности населения после п.3.7: {money(st.get('alt_sum_population'))} ₽."
                    )
                if st.get("alt_quarter_from_past_percent") is not None:
                    lines.append(
                        f"• Альтернативная квартальная сумма по п.3.7: {money(st.get('alt_quarter_from_past_percent'))} ₽."
                    )
                lines.append(f"• Итог после применения п.3.7: {money(st.get('min_after'))} ₽.")
        elif step == "SUBSCRIBERS_MIN":
            lines.append(
                f"• По абонентам: {money(st.get('subscriber_total'))} × 5 ₽ = {money(st.get('min_after'))} ₽."
            )
        if st.get("min_after") is not None:
            prev_min_after = float(st.get("min_after") or 0.0)
    return lines


def compute_percent_sum_q(
    contract_rate: float,
    annual_revenue: Optional[float],
    revenue_q: Optional[float],
    expenses_q: Optional[float],
) -> Tuple[Optional[Decimal], Dict[str, Any], List[str]]:
    notes: List[str] = []
    det: Dict[str, Any] = {"base_type": None, "base_q": None}

    base_q = None
    if revenue_q is not None:
        base_q = revenue_q
        det["base_type"] = "доходы (квартал)"
        notes.append("База для процента: доходы за квартал (введено пользователем).")
    elif annual_revenue is not None:
        base_q = annual_revenue / 4.0
        det["base_type"] = "доходы (год/4)"
        notes.append("База для процента: годовая выручка/доход разделён на 4 (если нет поквартальных данных).")
    elif expenses_q is not None:
        base_q = expenses_q
        det["base_type"] = "расходы (квартал)"
        notes.append("База для процента: расходы за квартал (ветка 100% госструктура / нет доходов).")
    else:
        notes.append("Не задана база для процента (нет доходов/выручки/расходов).")
        return None, det, notes

    det["base_q"] = base_q
    # По запросу: расчётные суммы по процентной ставке НЕ округляем.
    percent_sum_q = (Decimal(str(base_q)) * Decimal(str(contract_rate))) / Decimal("100")
    return percent_sum_q, det, notes


def lookup_min_sum(mins_df: pd.DataFrame, population: int, media_class: str) -> Optional[float]:
    sub = mins_df[
        (mins_df["Среда осуществления вещания (в эфире, по кабелю, одновременно в эфире и по кабелю)"]
         .astype(str).str.strip() == media_class)
    ].copy()
    if sub.empty:
        return None

    for _, r in sub.iterrows():
        lo = int(r["Численность населения на территории вещания, от (человек)"])
        hi = r["Численность населения на территории вещания, до (человек)"]
        hi_val = int(hi) if pd.notna(hi) else None
        if population >= lo and (hi_val is None or population <= hi_val):
            return float(r["Минимальная сумма авторского вознаграждения за квартал, рублей"])
    return None


def hour_coeff(hours_df: pd.DataFrame, hours_week: float) -> float:
    for _, r in hours_df.iterrows():
        lo = float(r["Количество часов вещания в неделю, от"])
        hi = float(r["Количество часов вещания в неделю, до"])
        if hours_week >= lo and hours_week <= hi:
            return float(r["Коэффициент к установленной минимальной сумме вознаграждения"])
    return 1.0


def discount_by_licenses(disc_df: pd.DataFrame, n_licenses: int) -> float:
    for _, r in disc_df.iterrows():
        lo = int(r["Минимальное количество вещательных лицензий одного пользователя"])
        hi = r["Максимальное количество вещательных лицензий одного пользователя"]
        hi_val = int(hi) if pd.notna(hi) else None
        if n_licenses >= lo and (hi_val is None or n_licenses <= hi_val):
            disc_pct = float(r["Размер скидки к совокупной минимальной сумме вознаграждения, процентов"])
            return 1.0 - disc_pct / 100.0
    return 1.0


def contract_period_coeff(period_df: pd.DataFrame, contract_quarter: int) -> float:
    for _, r in period_df.iterrows():
        lo = int(r["Отчетный период действия лицензионного договора, начиная с (номер квартала)"])
        hi = int(r["Отчетный период действия лицензионного договора, по (номер квартала включительно)"])
        if contract_quarter >= lo and contract_quarter <= hi:
            return float(r["Коэффициент к минимальной сумме вознаграждения в указанный период"])
    return 1.0


def compute_min_total(
    licenses: List[License],
    vars_xlsx: Path,
    annual_income_for_rules: Optional[float],
    contract_quarter: int,
    internet_resources: int,
    past_year_percent_paid: Optional[float],
    percent_sum_q: Optional[float],
    contract_media: str = "auto",
    use_small_income_branch: Optional[bool] = None,
    new_user_only: bool = False,
    assoc_member: bool = False,
    subscriber_total: Optional[int] = None,
    apply_license_count_discount: bool = True,
) -> Tuple[Optional[float], Dict[str, Any], List[str]]:
    """Блок C: расчёт минималки по таблицам + поправки (население/часы/интернет/скидка по числу лицензий — опционально)
    и затем «гильотина» как автоматическое ограничение минималки по сумме по проценту.

    ВАЖНО: коэффициенты льгот/стимулов (блок E) здесь НЕ применяются.
    """
    notes: List[str] = []
    details: Dict[str, Any] = {"steps": []}

    mins_df = pd.read_excel(vars_xlsx, sheet_name="Минимальные суммы по населению")
    disc_df = pd.read_excel(vars_xlsx, sheet_name="Скидки по количеству лицензий")
    hours_df = pd.read_excel(vars_xlsx, sheet_name="Коэффициенты по часам")
    params_df = pd.read_excel(vars_xlsx, sheet_name="Параметры для расчетов")

    def get_param_contains(substr: str, default: float) -> float:
        col = "Наименование параметра для расчета авторского вознаграждения"
        if col not in params_df.columns:
            return default
        sub = params_df[params_df[col].astype(str).str.contains(substr, case=False, na=False)]
        if sub.empty:
            return default
        return float(sub.iloc[0]["Значение параметра"])

    THRESH_SMALL = get_param_contains("Порог годового дохода", 1_500_000.0)
    SMALL_K = get_param_contains("Коэффициент уменьшения", 0.5)
    SMALL_MAX_Q = int(get_param_contains("Максимальное количество отчетных периодов применения половины", 8))
    INTERNET_PCT = get_param_contains("Дополнительный процент увеличения", 0.15)
    INTERNET_MIN_ADD = get_param_contains("Минимальное увеличение", 12500.0)
    GUILLOTINE_COEF = get_param_contains("Коэффициент гильотины", 1.10)
    if GUILLOTINE_COEF <= 1.0:
        # По п. 3.7 порог превышения — более чем на 10%.
        GUILLOTINE_COEF = 1.10

    pops_missing = [lic.license_id for lic in licenses if lic.population_total is None]
    if pops_missing:
        notes.append(f"Не найдена численность населения по лицензиям: {pops_missing}. Без населения минималка будет неполной.")

    # определяем среду договора на уровне агрегирования
    media_classes = [lic.media_class for lic in licenses]
    if "Одновременно в эфире и по кабелю" in media_classes:
        has_two_media = True
    else:
        has_air = any("эфир" in (lic.media_raw or "").lower() or "назем" in (lic.media_raw or "").lower() for lic in licenses)
        has_cable = any("кабель" in (lic.media_raw or "").lower() for lic in licenses)
        has_two_media = bool(has_air and has_cable)

    media_for_agg = "Одновременно в эфире и по кабелю" if has_two_media else "В эфире или по кабелю"

    contract_media = (contract_media or "auto").lower().strip()
    if contract_media in ("cable", "air"):
        has_two_media = False
        media_for_agg = "В эфире или по кабелю"
        notes.append("Среда договора принудительно задана как «В эфире или по кабелю» (эфир/кабель).")
    elif contract_media == "both":
        has_two_media = True
        media_for_agg = "Одновременно в эфире и по кабелю"
        notes.append("Среда договора принудительно задана как «Одновременно в эфире и по кабелю».")
    else:
        if any("универс" in str(lic.media_raw or "").lower() for lic in licenses):
            notes.append(
                "В режиме «Автоматически» универсальная лицензия считается как «Одновременно в эфире и по кабелю». "
                "Для расчёта как одной среды выберите среду договора вручную."
            )

    if use_small_income_branch is not None:
        small_branch = use_small_income_branch
    else:
        small_branch = bool(
            annual_income_for_rules is not None
            and annual_income_for_rules <= THRESH_SMALL
            and contract_quarter <= SMALL_MAX_Q
        )

    min_total = 0.0
    per_license_internet_present = any(int(getattr(lic, "internet_resources", 0) or 0) > 0 for lic in licenses)
    per_license_internet_applied = False

    # --- Ветка 3.4.2 / 3.5: минималка по абонентам ---
    if subscriber_total is not None:
        subs = int(subscriber_total)
        if subs < 0:
            return None, details, notes + ["Количество абонентов не может быть отрицательным."]
    
        min_total = float(subs) * 5.0
        details["steps"].append({"step": "SUBSCRIBERS_MIN", "subscriber_total": subs, "min_after": min_total})
        notes.append("Минимальная сумма рассчитана по абонентам: не менее 5 руб. за абонента (пп. 3.4.2 / 3.5).")
    
        # интернет-доплата (3.6) — оставляем как у тебя ниже, она применится дальше
    else:
        min_total = 0.0

    if subscriber_total is None:
    # C1/C2/C7
        if small_branch:
            N_sum = sum(int(lic.population_total) for lic in licenses if lic.population_total is not None)
            if N_sum <= 0:
                return None, details, notes + ["Нельзя применить ветку малого дохода: нет суммарной численности населения."]
            m = lookup_min_sum(mins_df, N_sum, media_for_agg)
            if m is None:
                return None, details, notes + ["Не найдена минималка в таблице по суммарной численности населения."]
            avg_hours = sum(float(lic.total_hours() or 0.0) for lic in licenses) / max(1, len(licenses))
            k_hours = hour_coeff(hours_df, avg_hours) if avg_hours < 126 else 1.0
            min_total = SMALL_K * m * k_hours
            details["steps"].append({
                "step": "C3",
                "N_sum": N_sum,
                "media": media_for_agg,
                "min_table": m,
                "k_small": SMALL_K,
                "avg_hours": avg_hours,
                "hour_coeff": k_hours,
                "min_after": min_total,
            })
            notes.append(
                "Включена ветка малого дохода: минималка по суммарной численности населения, "
                "затем коэффициент по среднему объёму вещания и ×0,5."
            )
        else:
            per_lic = []
            for lic in licenses:
                if lic.population_total is None:
                    continue
    
                media_for_min = lic.media_class
                if contract_media in ("cable", "air"):
                    media_for_min = "В эфире или по кабелю"
                elif contract_media == "both":
                    media_for_min = "Одновременно в эфире и по кабелю"
    
                m = lookup_min_sum(mins_df, int(lic.population_total), media_for_min)
                if m is None:
                    notes.append(f"Не найдена минималка по таблице для лицензии {lic.license_id} (население={lic.population_total}, среда={media_for_min}).")
                    continue
    
                hrs = lic.total_hours()
                coeff = 1.0
                if hrs < 126:
                    coeff = hour_coeff(hours_df, hrs)
    
                m2 = m * coeff
                if coeff < 1.0:
                    notes.append(
                        f"Коэффициент по объёму вещания применён для лицензии {lic.license_id}: "
                        f"{hrs:g} ч/нед → {str(coeff).replace('.', ',')}."
                    )
                lic_res = int(getattr(lic, "internet_resources", 0) or 0)
                if lic_res > 0:
                    add_per_lic = max(INTERNET_PCT * m2, INTERNET_MIN_ADD)
                    delta_lic = add_per_lic * lic_res
                    m2 += delta_lic
                    per_license_internet_applied = True
                    delta_txt = f"{int(round(delta_lic)):,}".replace(",", " ")
                    notes.append(
                        f"Интернет-доплата применена по лицензии {lic.license_id}: "
                        f"{lic_res} ресурс(ов), +{delta_txt} ₽."
                    )
                per_lic.append({"license_id": lic.license_id, "population": lic.population_total, "media": media_for_min, "min_table": m, "hours_week": hrs, "hour_coeff": coeff, "min_after": m2})
                min_total += m2
    
            details["steps"].append({"step": "C1+C2(+C7)", "per_license": per_lic, "min_after": min_total})

    # C4 скидка по числу лицензий
    n_lic = len(licenses)
    if apply_license_count_discount and n_lic > 3:
        k = discount_by_licenses(disc_df, n_lic)
        min_total *= k
        floor_single_medium = None
        try:
            one_medium = mins_df[
                mins_df["Среда осуществления вещания (в эфире, по кабелю, одновременно в эфире и по кабелю)"]
                .astype(str).str.strip() == "В эфире или по кабелю"
            ].copy()
            if not one_medium.empty:
                floor_single_medium = 0.5 * float(
                    one_medium["Минимальная сумма авторского вознаграждения за квартал, рублей"].min()
                )
        except Exception:
            floor_single_medium = None

        if floor_single_medium is not None and min_total < floor_single_medium:
            min_total = float(floor_single_medium)
            notes.append(
                "После скидки по числу лицензий применено нижнее ограничение: не ниже 1/2 минимальной суммы для одной лицензии в одной среде вещания."
            )

        details["steps"].append({
            "step": "C4",
            "n_licenses": n_lic,
            "coeff": k,
            "floor_single_medium_half": floor_single_medium,
            "min_after": min_total
        })
        notes.append(f"Применена скидка по числу лицензий (кол-во ВЛ={n_lic}).")

    # C6 интернет-доплата (контрактный режим — используется только когда нет разметки по лицензиям)
    if (not per_license_internet_applied) and per_license_internet_present:
        resources_total = sum(int(getattr(lic, "internet_resources", 0) or 0) for lic in licenses)
        if resources_total > 0:
            add_per = max(INTERNET_PCT * min_total, INTERNET_MIN_ADD)
            delta = add_per * resources_total
            min_total += delta
            details["steps"].append({"step": "C6", "resources": resources_total, "add_per_resource": add_per, "delta": delta, "min_after": min_total})
            notes.append("Добавлена доплата за интернет-вещание по суммарному количеству ресурсов (перелицензионная ветка недоступна для текущего режима расчёта).")
    elif internet_resources and internet_resources > 0:
        add_per = max(INTERNET_PCT * min_total, INTERNET_MIN_ADD)
        delta = add_per * internet_resources
        min_total += delta
        details["steps"].append({"step": "C6", "resources": internet_resources, "add_per_resource": add_per, "delta": delta, "min_after": min_total})
        notes.append("Добавлена доплата за интернет-вещание (+15%, но не менее 12 500 за ресурс).")

    # «Гильотина» по п. 3.7:
    # 1) если минималка > 1.10 * сумма по проценту, переходим на минималку п. 3.1
    #    по суммарной численности населения (независимо от числа лицензий);
    # 2) для ветки 3.2 (малый доход) при повторном превышении применяем 1/4
    #    от фактических выплат по проценту за предшествующий год (если переданы).
    if percent_sum_q is not None:
        min_before = float(min_total)
        threshold = float(GUILLOTINE_COEF) * float(percent_sum_q)
        guillotine_triggered = bool(min_before > threshold)
        min_after = min_before
        alt1 = None
        alt2 = None

        if guillotine_triggered:
            n_sum_all = sum(int(lic.population_total) for lic in licenses if lic.population_total is not None)
            if n_sum_all > 0:
                alt1 = lookup_min_sum(mins_df, n_sum_all, media_for_agg)
                if alt1 is not None:
                    avg_hours = sum(float(lic.total_hours()) for lic in licenses) / max(1, len(licenses))
                    k_hours = hour_coeff(hours_df, avg_hours) if avg_hours < 126 else 1.0
                    alt1 = float(alt1) * float(k_hours)
                    min_after = float(alt1)
                    notes.append(
                        "Гильотина (п.3.7): применена минималка по п.3.1 от суммарной численности населения."
                    )

            # Дополнительная ступень: если даже пересчитанная минималка остаётся выше порога,
            # применяем сумму по процентной ставке (для малого дохода — при наличии истории используем 1/4 прошлых выплат).
            if min_after > threshold:
                if small_branch and contract_quarter <= SMALL_MAX_Q and past_year_percent_paid is not None and past_year_percent_paid > 0:
                    alt2 = float(past_year_percent_paid) / 4.0
                    notes.append(
                        "Гильотина (п.3.7, абз.2): применена 1/4 от фактических выплат по проценту за предыдущий год."
                    )
                else:
                    alt2 = float(percent_sum_q)
                    notes.append(
                        "Гильотина (п.3.7): после проверки по суммарной численности населения применена сумма по процентной ставке."
                    )
                min_after = float(alt2)

        min_total = float(min_after)

        details["steps"].append({
            "step": "GUILLOTINE",
            "triggered": guillotine_triggered,
            "min_before": round(min_before, 2),
            "s_percent": round(float(percent_sum_q), 2),
            "threshold_coef": float(GUILLOTINE_COEF),
            "threshold": round(threshold, 2),
            "alt_sum_population": round(float(alt1), 2) if alt1 is not None else None,
            "alt_quarter_from_past_percent": round(float(alt2), 2) if alt2 is not None else None,
            "min_after": round(float(min_total), 2),
        })

    return round_rub(min_total), details, notes


# --------------------------- reporting ---------------------------

def money(x: Optional[float]) -> str:
    if x is None:
        return "—"
    return f"{x:,.2f}".replace(",", " ").replace(".00", "")


def money_precise(x: Optional[Any]) -> str:
    """Без округления: для расчётных сумм по процентной ставке."""
    if x is None:
        return "—"
    d = Decimal(str(x))
    s = format(d, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    sign = "-" if s.startswith("-") else ""
    if sign:
        s = s[1:]
    int_part, dot, frac_part = s.partition(".")
    int_fmt = f"{int(int_part or '0'):,}".replace(",", " ")
    return f"{sign}{int_fmt}{('.' + frac_part) if dot and frac_part else ''}"


def format_report(
    inn: str,
    year: Optional[int],
    annual_revenue: Optional[float],
    revenue_q: Optional[float],
    expenses_q: Optional[float],
    internet_resources: int,
    contract_quarter: int,
    new_user_only: bool,
    user_discount_label: str,
    fixed_fee_eligible: bool,
    manual_retransmission: bool,
    rate_mode: str,
    actual_usage_share: Optional[float],
    actual_usage_applied: bool,
    actual_usage_category: Optional[str],
    licenses: List[License],
    contract_rate: float,
    base_percent_sum_q: Optional[float],
    base_min_total: Optional[float],
    final_percent_sum_q: Optional[float],
    final_min_total: Optional[float],
    min_with_coeff_y1: Optional[float],
    min_with_coeff_y2: Optional[float],
    addendum_percent_sum_q: Optional[float],
    addendum_min_total: Optional[float],
    addendum_min_with_coeff_y1: Optional[float],
    addendum_min_with_coeff_y2: Optional[float],
    addendum_base_before_coeffs: Optional[float],
    addendum_base_after_common_coeffs: Optional[float],
    k_license_count: float,
    k42: float,
    guillotine_triggered: bool,
    addendum_limited_8q: bool,
    addendum_percent_based: bool,
    applied_coeffs: List[str],
    volume_coeff_notes: List[str],
    minimum_breakdown: List[str],
    notes: List[str],
    needs: List[str],
    society: str = "РАО",
) -> str:
    lines: List[str] = []
    sep = "────────────────────────────────────────"
    usage_obj_gen = "фонограмм" if str(society).strip().upper() == "ВОИС" else "произведений"

    org_name = licenses[0].org_name if licenses else "Не найдено (нет записей в РКН)"
    lines.append(org_name)
    lines.append(f"ИНН: {inn}")
    lines.append("")
    lines.append(sep)
    lines.append("ИСХОДНЫЕ ДАННЫЕ")
    lines.append(sep)

    if revenue_q is not None:
        lines.append(f"Доходы за квартал (введено): {money(revenue_q)} ₽.")
    elif annual_revenue is not None:
        lines.append(f"Выручка/доход за {year or 'год'} (введено): {money(annual_revenue)} ₽.")
        lines.append("База квартала: год/4 (если нет поквартальных данных).")
    elif expenses_q is not None:
        lines.append(f"Расходы за квартал (введено): {money(expenses_q)} ₽ (ветка 100% госструктура / нет доходов).")
    else:
        lines.append("Финансовая база для расчёта процента: НЕ ЗАДАНА.")
    lines.append("")

    lines.append(f"Ранее работал с {society}: {'да' if (not new_user_only) else 'нет'}.")
    lines.append(f"Альтернативный коэффициент (п.4.2): {user_discount_label or 'не применяется'}.")
    if fixed_fee_eligible:
        lines.append("Пользователь относится к категории, для которой может применяться фиксированный размер вознаграждения.")
    if manual_retransmission:
        lines.append("Есть ретрансляция стороннего телеканала: для спорных случаев ставку по лицензии можно задать вручную.")
    lines.append("")

    lines.append(f"Вещательные лицензии (по таблице РКН): {len(licenses)} шт.")
    for lic in licenses:
        pop = lic.population_total
        pop_str = f"{pop:,}".replace(",", " ") if pop is not None else "не найдено"
        smi_join = " / ".join([ch.name for ch in lic.channels if getattr(ch, "name", None)]) or "—"
        lines.append("")
        lines.append(f"▪ Лицензия: {lic.license_id}")
        lines.append(f"  Среда: {lic.media_raw} → {lic.media_class}")
        lines.append(f"  Население: {pop_str}")
        lines.append(f"  СМИ/каналы: {smi_join}")
        lines.append(f"  РКН: {lic.rkn_url}")
        if int(getattr(lic, "internet_resources", 0) or 0) > 0:
            lines.append(f"  Интернет-ресурсов по лицензии: {int(getattr(lic, 'internet_resources', 0))}")
        if getattr(lic, "internet_sites", None):
            lines.append(f"  Интернет-сайт(ы) по лицензии: {'; '.join([str(x) for x in lic.internet_sites])}")
        if lic.population_notes:
            for n in lic.population_notes[:2]:
                lines.append(f"  ⚑ Примечание: {n}")

        lic_rate, _ = compute_license_rate(lic)
        if lic.rate_override is not None:
            lines.append(f"  Ставка по лицензии: {lic_rate:.1f}% (задана вручную)")
        else:
            lines.append(f"  Ставка по лицензии: {lic_rate:.1f}%")
        for ch in lic.channels[:10]:
            hrs = ch.hours_week if ch.hours_week is not None else 168.0
            ch_rate, ch_notes = ch.avg_rate()
            lines.append(f"  • Канал/СМИ: {ch.name}")
            lines.append(f"    Часы/нед: {hrs:g}")
            lines.append(f"    Ставка канала: {round_rate(ch_rate):.1f}%")
            for hn in (ch.hours_notes or [])[:2]:
                lines.append(f"    ◦ {hn}")
            for tn in ch_notes[:2]:
                lines.append(f"    ◦ {tn}")
            if ch.topics:
                for t in ch.topics:
                    share = f"{t.share_pct:g}%" if t.share_pct is not None else "без доли"
                    lines.append(f"    ◦ Тематика: {t.topic_raw} ({share}) → {t.rate_pct:.1f}%")

    lines.append("")
    if minimum_breakdown:
        lines.append("Расчёт минимальной суммы:")
        for line in minimum_breakdown:
            lines.append(line)
        lines.append("")
    lines.append(f"Процентная ставка по договору (взвешенная по часам×населению): {contract_rate:.1f}%.")
    lines.append(f"Расчётная сумма по договору за квартал: {money_precise(base_percent_sum_q)} ₽.")
    if rate_mode == "actual_share":
        if actual_usage_applied:
            if actual_usage_category == "по каждому СМИ":
                lines.append(
                    f"Режим ставки: по фактической доле использования {usage_obj_gen} по каждому СМИ "
                    "(введено вручную по выбранным лицензиям/каналам)."
                )
            else:
                share_txt = f"{float(actual_usage_share):g}%" if actual_usage_share is not None else "—"
                cat_txt = actual_usage_category or "—"
                lines.append(
                    f"Режим ставки: по фактической доле использования {usage_obj_gen} "
                    f"(доля: {share_txt}, категория: {cat_txt})."
                )
        else:
            lines.append(
                f"Режим ставки по фактической доле использования {usage_obj_gen} запрошен, "
                "но условия пересмотра не подтверждены; применена ставка по лицензиям."
            )
    lines.append("")
    lines.append(sep)
    lines.append("УСЛОВИЯ ПО ДОГОВОРУ")
    lines.append(sep)

    calc_explain_notes: List[str] = []
    explain_markers = (
        "скидка по числу лицензий",
        "ветка малого дохода",
        "гильотина",
        "доплата за интернет-вещание",
        "нижнее ограничение",
        "среда договора принудительно",
        "универсальная лицензия считается",
    )
    for n in notes:
        nl = str(n).strip()
        low = nl.lower()
        if any(m in low for m in explain_markers):
            calc_explain_notes.append(nl)

    if calc_explain_notes:
        lines.append("Пояснения к расчёту минимальной суммы:")
        for n in calc_explain_notes:
            lines.append(f"• {n}")
        lines.append("")

    lines.append("Базовые условия по договору:")
    if volume_coeff_notes:
        lines.append("• Применён коэффициент по объёму вещания:")
        for vc in volume_coeff_notes:
            lines.append(f"  - {vc}")
    if fixed_fee_eligible:
        base_fix = None
        if base_percent_sum_q is not None or base_min_total is not None:
            base_fix = max(float(base_percent_sum_q or 0.0), float(base_min_total or 0.0))
        lines.append(f"• Фиксированный размер вознаграждения: {money(base_fix)} ₽ за квартал.")
    else:
        lines.append(f"• Расчётная сумма по договору за квартал: {money_precise(base_percent_sum_q)} ₽.")
        if base_min_total is not None:
            lines.append(f"• Условие договора: {contract_rate:.1f}% от дохода за квартал, но не менее {money(base_min_total)} ₽ за квартал.")
    lines.append("")

    lines.append("Условия по договору с применением коэффициентов:")
    if applied_coeffs:
        lines.append(f"• Применённые коэффициенты: {', '.join(applied_coeffs)}.")
    else:
        lines.append("• Применённые коэффициенты: не применяются.")
    if fixed_fee_eligible:
        final_fix = None
        if final_percent_sum_q is not None or final_min_total is not None:
            final_fix = max(float(final_percent_sum_q or 0.0), float(final_min_total or 0.0))
        lines.append(f"• Фиксированный размер вознаграждения: {money(final_fix)} ₽ за квартал.")
    else:
        lines.append(f"• Расчётная сумма по договору за квартал: {money_precise(final_percent_sum_q)} ₽.")
        if new_user_only and (min_with_coeff_y1 is not None or min_with_coeff_y2 is not None):
            if min_with_coeff_y1 is not None:
                lines.append(
                    f"• Условие договора на первый год (1–4 отчётные периоды): "
                    f"{contract_rate:.1f}% от дохода за квартал, но не менее {money(min_with_coeff_y1)} ₽ за квартал."
                )
            if min_with_coeff_y2 is not None:
                lines.append(
                    f"• Условие договора на второй год (5–8 отчётные периоды): "
                    f"{contract_rate:.1f}% от дохода за квартал, но не менее {money(min_with_coeff_y2)} ₽ за квартал."
                )
            if final_min_total is not None:
                lines.append(
                    f"• Для текущего отчётного периода №{contract_quarter}: "
                    f"{contract_rate:.1f}% от дохода за квартал, но не менее {money(final_min_total)} ₽ за квартал."
                )
        elif final_min_total is not None:
            lines.append(
                f"• Условие договора: {contract_rate:.1f}% от дохода за квартал, "
                f"но не менее {money(final_min_total)} ₽ за квартал."
            )
    lines.append("")

    if (addendum_min_total is not None or addendum_percent_sum_q is not None):
        lines.append("Условия по дополнительному соглашению к договору:")
        addendum_tail = ""
        if addendum_limited_8q:
            addendum_tail += " (на период не превышающий 8 отчётных периодов)"
        if addendum_percent_based:
            addendum_tail += " (по проценту)"
        if fixed_fee_eligible:
            add_fix = max(float(addendum_percent_sum_q or 0.0), float(addendum_min_total or 0.0))
            lines.append(f"• Фиксированный размер вознаграждения: {money(add_fix)} ₽ за квартал.")
        else:
            lines.append(f"• Расчётная сумма по договору за квартал: {money_precise(addendum_percent_sum_q)} ₽.")
            if (
                addendum_base_before_coeffs is not None
                and addendum_base_after_common_coeffs is not None
                and (
                    abs(float(k_license_count or 1.0) - 1.0) > 1e-9
                    or abs(float(k42 or 1.0) - 1.0) > 1e-9
                )
            ):
                coeff_parts: List[str] = []
                if abs(float(k_license_count or 1.0) - 1.0) > 1e-9:
                    coeff_parts.append(str(float(k_license_count)).replace(".", ","))
                if abs(float(k42 or 1.0) - 1.0) > 1e-9:
                    coeff_parts.append(str(float(k42)).replace(".", ","))
                coeff_formula = " × ".join(coeff_parts) if coeff_parts else "1"
                lines.append(
                    f"• Минимальная сумма доп. соглашения после коэффициентов: {money(addendum_base_before_coeffs)} ₽ × {coeff_formula} = {money(addendum_base_after_common_coeffs)} ₽."
                )
            if new_user_only and (addendum_min_with_coeff_y1 is not None or addendum_min_with_coeff_y2 is not None):
                if addendum_min_with_coeff_y1 is not None:
                    lines.append(
                        f"• Условие доп. соглашения на первый год (1–4 отчётные периоды): "
                        f"{contract_rate:.1f}% от дохода за квартал, но не менее {money(addendum_min_with_coeff_y1)} ₽ за квартал{addendum_tail}."
                    )
                if addendum_min_with_coeff_y2 is not None:
                    lines.append(
                        f"• Условие доп. соглашения на второй год (5–8 отчётные периоды): "
                        f"{contract_rate:.1f}% от дохода за квартал, но не менее {money(addendum_min_with_coeff_y2)} ₽ за квартал{addendum_tail}."
                    )
                if addendum_min_total is not None:
                    lines.append(
                        f"• Для текущего отчётного периода №{contract_quarter}: "
                        f"{contract_rate:.1f}% от дохода за квартал, но не менее {money(addendum_min_total)} ₽ за квартал{addendum_tail}."
                    )
            elif addendum_min_total is not None:
                lines.append(
                    f"• Условие доп. соглашения: {contract_rate:.1f}% от дохода за квартал, "
                    f"но не менее {money(addendum_min_total)} ₽ за квартал{addendum_tail}."
                )
        lines.append("")
    else:
        lines.append("Условия по дополнительному соглашению к договору: не применяются.")
        lines.append("")

    per_license_resources = {
        str(lic.license_id): int(getattr(lic, "internet_resources", 0) or 0)
        for lic in (licenses or [])
        if int(getattr(lic, "internet_resources", 0) or 0) > 0
    }
    if per_license_resources:
        pairs = "; ".join([f"{k}: {v}" for k, v in per_license_resources.items()])
        lines.append(f"Интернет-вещание по лицензиям: {pairs}.")
        lines.append("")
    elif internet_resources:
        lines.append(f"Интернет-вещание: указано ресурсов — {internet_resources} (применена доплата по правилам).")
        lines.append("")

    if needs:
        lines.append("Нужно уточнить/проверить:")
        for x in needs:
            lines.append(f"• {x}")
        lines.append("")

    if notes:
        lines.append(sep)
        lines.append("ТЕХНИЧЕСКИЕ ПРИМЕЧАНИЯ И ДОПУЩЕНИЯ")
        lines.append(sep)
        for n in notes:
            lines.append(f"• {n}")
        lines.append("")

    return "\n".join(lines)


# --------------------------- main (non-interactive) ---------------------------

def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--inn", required=True, help="ИНН (10/12 цифр)")
    ap.add_argument("--year", type=int, default=None, help="Год (для подписи в отчёте)")

    ap.add_argument("--annual_revenue", type=float, default=None, help="Годовая выручка/доход")
    ap.add_argument("--revenue_q", type=float, default=None, help="Доходы за квартал")
    ap.add_argument("--expenses_q", type=float, default=None, help="Расходы за квартал (ветка госструктуры)")

    ap.add_argument("--internet_resources", type=int, default=0)
    ap.add_argument("--contract_quarter", type=int, default=1)
    ap.add_argument("--contract_media", type=str, default="auto", choices=["auto", "cable", "air", "both"])

    ap.add_argument("--new_user", action="store_true")
    ap.add_argument("--assoc_member", action="store_true")
    ap.add_argument("--is_package_contract", action="store_true")
    ap.add_argument("--is_new_user_and_other_use_contract", action="store_true")
    ap.add_argument("--fixed_fee_eligible", action="store_true")
    ap.add_argument("--first_year_entity", action="store_true")
    ap.add_argument("--manual_retransmission", action="store_true")

    ap.add_argument("--only_license", type=str, default=None)
    ap.add_argument("--licenses", action="append", default=None, help="Номера лицензий (можно несколько, через запятую)")

    ap.add_argument("--rkn_xlsx", type=str, default="Таблица РКН slim.xlsx")
    ap.add_argument("--vars_xlsx", type=str, default="Переменные из ставок.xlsx")

    ap.add_argument("--force_small_income", action="store_true")
    ap.add_argument("--no_small_income", action="store_true")

    ap.add_argument("--population_override", type=int, default=None)
    ap.add_argument("--population_by_license", action="append", default=None, help="Переопределение населения: ЛИЦЕНЗИЯ=НАСЕЛЕНИЕ")
    ap.add_argument("--internet_resources_by_license", action="append", default=None, help="Интернет-ресурсы по лицензии: ЛИЦЕНЗИЯ=КОЛИЧЕСТВО")
    ap.add_argument("--internet_sites_by_license", action="append", default=None, help="Интернет-сайты по лицензии: ЛИЦЕНЗИЯ=URL1|URL2")
    ap.add_argument("--skip_runtime_population_normalization", action="store_true")
    ap.add_argument("--rate_by_license", action="append", default=None, help="Ручная ставка по лицензии: ЛИЦЕНЗИЯ=СТАВКА")
    ap.add_argument("--rate_mode", type=str, default="license", choices=["license", "actual_share"], help="Режим расчёта ставки по договору")
    ap.add_argument("--actual_share_by_channel", action="append", default=None, help="Фактическая доля по СМИ: ЛИЦЕНЗИЯ|СМИ=87.5 или ЛИЦЕНЗИЯ|1=87.5")
    ap.add_argument("--actual_usage_share", type=float, default=None, help="Фактическая доля использования произведений, %%")
    ap.add_argument("--actual_usage_confirmed", action="store_true", help="Подтверждены данные минимум за 4 полных отчётных периода")
    ap.add_argument("--actual_usage_change_ge20", action="store_true", help="Изменение доли использования >= 20%%")
    ap.add_argument("--actual_usage_periods", type=int, default=0, help="Количество полных отчётных периодов подтверждённой фактической доли")
    ap.add_argument("--subscriber_total", type=int, default=None, help="Суммарное количество абонентов (ветка 3.4.2/3.5)")
    ap.add_argument("--past_year_percent_paid", type=float, default=None, help="Фактические выплаты по проценту за предшествующий год (для п.3.7)")

    ap.add_argument("--non_interactive", action="store_true", help="для сайта всегда ставим этот флаг")
    ap.add_argument("--society", type=str, default="РАО", help="Наименование организации в отчёте (РАО/ВОИС)")

    args = ap.parse_args(argv)

    p = Progress(enabled=not bool(args.non_interactive))

    try:
        inn = parse_inn(args.inn)
    except Exception as e:
        print(f"Ошибка: {e}")
        return 2

    base_dir = Path(__file__).resolve().parent
    project_dir = Path(__file__).resolve().parents[3]
    tv_data_dir = project_dir / "calculators" / "rao_tv" / "data"
    rkn_xlsx = Path(args.rkn_xlsx)
    vars_xlsx = Path(args.vars_xlsx)
    if not rkn_xlsx.is_absolute():
        if (tv_data_dir / rkn_xlsx).exists():
            rkn_xlsx = tv_data_dir / rkn_xlsx
        elif (project_dir / rkn_xlsx).exists():
            rkn_xlsx = project_dir / rkn_xlsx
        else:
            rkn_xlsx = base_dir / rkn_xlsx
    if not vars_xlsx.is_absolute():
        if (tv_data_dir / vars_xlsx).exists():
            vars_xlsx = tv_data_dir / vars_xlsx
        elif (project_dir / vars_xlsx).exists():
            vars_xlsx = project_dir / vars_xlsx
        else:
            vars_xlsx = base_dir / vars_xlsx

    if not rkn_xlsx.exists():
        for name in (
            "Таблица РКН slim.xlsx",
            "Таблица РКН очищенная.xlsx",
            "Таблица РКН.xlsx",
            "Таблица РКН (2).xlsx",
        ):
            alt = rkn_xlsx.with_name(name)
            if alt.exists():
                rkn_xlsx = alt
                break
            alt2 = tv_data_dir / name
            if alt2.exists():
                rkn_xlsx = alt2
                break

    if not rkn_xlsx.exists():
        print(f"Ошибка: не найден файл РКН: {rkn_xlsx}")
        return 2
    if not vars_xlsx.exists():
        print(f"Ошибка: не найден файл ставок: {vars_xlsx}")
        return 2

    p.tick("читаю РКН и собираю лицензии")
    licenses, load_notes = load_licenses_by_inn(
        rkn_xlsx,
        inn,
        vars_xlsx,
        runtime_population_normalization=not bool(args.skip_runtime_population_normalization),
        strict_topic_match=True,
    )

    selected_ids = parse_license_list(args.licenses, args.only_license)
    if selected_ids:
        selected_set = {str(x).strip() for x in selected_ids}
        licenses = [x for x in licenses if str(x.license_id).strip() in selected_set]
        if not licenses:
            print(f"Ошибка: не найдены лицензии {sorted(selected_set)} у этого ИНН в таблице РКН.")
            return 2

    pop_by_license = parse_population_by_license(args.population_by_license)
    rate_by_license = parse_rate_by_license(args.rate_by_license)
    actual_share_by_channel = parse_actual_share_by_channel(args.actual_share_by_channel)
    internet_by_license = parse_internet_resources_by_license(args.internet_resources_by_license)
    internet_sites_by_license = parse_internet_sites_by_license(args.internet_sites_by_license)

    if args.population_override is not None:
        po = int(args.population_override)
        for lic in licenses:
            if str(lic.license_id).strip() in pop_by_license:
                continue
            old = lic.population_total
            lic.population_total = po
            note = f"Переопределено пользователем: {po}" + (f" (РКН: {old})" if old is not None else "")
            lic.population_notes.append(note)

    if pop_by_license:
        for lic in licenses:
            key = str(lic.license_id).strip()
            if key not in pop_by_license:
                continue
            po = int(pop_by_license[key])
            old = lic.population_total
            lic.population_total = po
            note = f"Переопределено пользователем по лицензии: {po}" + (f" (РКН: {old})" if old is not None else "")
            lic.population_notes.append(note)

    if rate_by_license:
        for lic in licenses:
            key = str(lic.license_id).strip()
            if key not in rate_by_license:
                continue
            lic.rate_override = float(rate_by_license[key])

    if internet_by_license:
        for lic in licenses:
            key = str(lic.license_id).strip()
            lic.internet_resources = int(internet_by_license.get(key, 0) or 0)
    if internet_sites_by_license:
        for lic in licenses:
            key = str(lic.license_id).strip()
            lic.internet_sites = list(internet_sites_by_license.get(key, []) or [])

    internet_resources_effective = int(args.internet_resources or 0)
    if internet_by_license:
        internet_resources_effective = 0

    needs: List[str] = []
    notes: List[str] = []
    notes.extend(load_notes)

    if not licenses:
        print("Ошибка: нет данных по ИНН в таблице РКН.")
        return 2

    if all(lic.population_total is None for lic in licenses):
        needs.append("В РКН-таблице не заполнено население. Нужно взять численность населения территории вещания из карточек РКН.")
    if args.manual_retransmission:
        needs.append("Проверить тематики и параметры ретранслируемого телеканала вручную по лицензии учредителя канала.")

    actual_usage_applied = False
    actual_usage_category: Optional[str] = None
    usage_obj_gen = "фонограмм" if str(args.society or "").strip().upper() == "ВОИС" else "произведений"

    factual_conditions_ok = True
    if args.rate_mode == "actual_share":
        periods_ok = int(args.actual_usage_periods or 0) >= 4
        confirmed_ok = bool(args.actual_usage_confirmed)
        delta_ok = bool(args.actual_usage_change_ge20)
        factual_conditions_ok = bool(periods_ok and confirmed_ok and delta_ok)
        if not factual_conditions_ok:
            needs.append(
                "Для пересмотра ставки по фактической доле нужны подтверждённые данные за 4 полных отчётных периода и изменение доли не менее 20%."
            )
            notes.append("Условия пересмотра ставки по фактической доле не подтверждены; применена ставка по лицензиям.")
        elif actual_share_by_channel:
            notes.extend(
                apply_actual_share_by_channel(
                    licenses=licenses,
                    vars_xlsx=vars_xlsx,
                    actual_share_by_channel=actual_share_by_channel,
                    usage_obj_gen=usage_obj_gen,
                )
            )
            actual_usage_applied = True
            actual_usage_category = "по каждому СМИ"
        elif args.actual_usage_share is None:
            needs.append(f"Для режима пересмотра ставки укажите фактическую долю использования {usage_obj_gen}, %.")
            notes.append("Режим фактической доли выбран, но доля не задана; применена ставка по лицензиям.")

    p.tick("считаю процентную ставку по договору")
    contract_rate, _ = compute_contract_rate(licenses)

    if args.rate_mode == "actual_share" and factual_conditions_ok and not actual_share_by_channel and args.actual_usage_share is not None:
        factual_rate, factual_cat, factual_note = contract_rate_from_actual_usage_share(
            vars_xlsx, float(args.actual_usage_share)
        )
        if factual_rate is not None:
            contract_rate = float(factual_rate)
            actual_usage_applied = True
            actual_usage_category = factual_cat
            notes.append(f"Ставка по договору пересчитана по фактической доле использования {usage_obj_gen}: {factual_note}")
        else:
            needs.append(f"Не удалось рассчитать ставку по фактической доле использования {usage_obj_gen}.")
            notes.append(f"Пересчёт по фактической доле не выполнен: {factual_note}")

    p.tick("считаю сумму по проценту за квартал")
    percent_sum_q, _, percent_notes = compute_percent_sum_q(
        contract_rate=contract_rate,
        annual_revenue=args.annual_revenue,
        revenue_q=args.revenue_q,
        expenses_q=args.expenses_q,
    )
    notes.extend(percent_notes)
    base_percent_sum_q = percent_sum_q

    if percent_sum_q is None:
        needs.append("Нужна финансовая база: годовая выручка/доход или доходы за квартал или расходы за квартал (для ветки госструктуры).")

    annual_income_for_rules = None
    if args.annual_revenue is not None:
        annual_income_for_rules = float(args.annual_revenue)
    elif args.revenue_q is not None:
        annual_income_for_rules = float(args.revenue_q) * 4.0

    if args.force_small_income and args.no_small_income:
        print("Ошибка: нельзя одновременно --force_small_income и --no_small_income")
        return 2

    use_small_income = None
    if args.force_small_income:
        use_small_income = True
    elif args.no_small_income:
        use_small_income = False

    p.tick("считаю минимальную сумму")
    p.tick("считаю минимальную сумму за квартал")
    min_total, min_details, min_notes = compute_min_total(
        licenses=licenses,
        vars_xlsx=vars_xlsx,
        annual_income_for_rules=annual_income_for_rules,
        contract_quarter=args.contract_quarter,
        internet_resources=internet_resources_effective,
        past_year_percent_paid=args.past_year_percent_paid,
        percent_sum_q=percent_sum_q,
        contract_media=args.contract_media,
        use_small_income_branch=use_small_income,
        new_user_only=bool(args.new_user),
        assoc_member=bool(args.assoc_member),
        subscriber_total=args.subscriber_total,
        apply_license_count_discount=False,

    )
    notes.extend(min_notes)
    base_min_total = min_total
    minimum_breakdown = build_minimum_breakdown_lines(min_details)

    small_income_applied = any(
        str(st.get("step") or "").upper() == "C3"
        for st in (min_details.get("steps") or [])
    )
    volume_coeff_notes: List[str] = []
    for st in (min_details.get("steps") or []):
        if str(st.get("step") or "").upper() != "C1+C2(+C7)":
            continue
        for row in (st.get("per_license") or []):
            try:
                hc = float(row.get("hour_coeff", 1.0))
            except Exception:
                hc = 1.0
            if hc >= 1.0:
                continue
            lid = str(row.get("license_id") or "—")
            hrs = float(row.get("hours_week") or 0.0)
            volume_coeff_notes.append(
                f"{lid}: {hrs:g} ч/нед → коэффициент {str(hc).replace('.', ',')}"
            )

    # Блок E / раздел 4:
    # - п.4.2: применяется только ОДИН коэффициент из доступных категорий (к % и к минималке);
    # - п.4.3: коэффициент нового пользователя по периодам 1-4 / 5-8 (только к минималке).
    try:
        down_df = pd.read_excel(vars_xlsx, sheet_name="Понижающие коэффициенты")
    except Exception:
        down_df = None

    def down_coeff(contains: str, default: float = 1.0) -> float:
        if down_df is None:
            return default
        col = "Условие применения понижающего коэффициента (описательное обозначение)"
        if col not in down_df.columns:
            return default
        sub = down_df[down_df[col].astype(str).str.contains(contains, case=False, na=False)]
        if sub.empty:
            return default
        return float(sub.iloc[0]["Понижающий коэффициент к минимальной сумме вознаграждения"])

    guillotine_step = None
    for st in (min_details.get("steps") or []):
        if str(st.get("step") or "").upper() == "GUILLOTINE":
            guillotine_step = st
            break
    guillotine_triggered = bool((guillotine_step or {}).get("triggered"))
    min_before_guillotine = float((guillotine_step or {}).get("min_before")) if (guillotine_step and guillotine_step.get("min_before") is not None) else float(min_total or 0.0)

    # п.4.2 — выбираем только один коэффициент (максимальную льготу = минимальный k)
    coeff_42_choices: List[Tuple[str, float]] = []
    if bool(args.assoc_member):
        coeff_42_choices.append(("член отраслевой ассоциации", down_coeff("ассоциаци", 0.85)))
    if bool(args.is_package_contract):
        coeff_42_choices.append(("пакетное заключение", down_coeff("пакет", 0.9)))
    if bool(args.is_new_user_and_other_use_contract):
        coeff_42_choices.append(("несколько категорий использования", down_coeff("нескольк", 0.95)))

    k42 = 1.0
    user_discount_label = "не применяется"
    if coeff_42_choices:
        user_discount_desc, k42 = min(coeff_42_choices, key=lambda x: x[1])
        user_discount_label = f"{str(k42).replace('.', ',')} ({user_discount_desc})" if k42 != 1.0 else "не применяется"
        if k42 != 1.0:
            notes.append(f"п.4.2: применён коэффициент {k42:g} ({user_discount_desc}) к минимальной сумме.")

    # п.4.3 — новый пользователь, коэффициент по периодам (только минималка).
    if bool(args.new_user):
        q = int(args.contract_quarter or 1)
        if 1 <= q <= 4:
            notes.append("п.4.3: для текущего периода применяются условия первого года (коэффициент 0,75 к минимальной сумме).")
        elif 5 <= q <= 8:
            notes.append("п.4.3: для текущего периода применяются условия второго года (коэффициент 0,88 к минимальной сумме).")
        else:
            notes.append("п.4.3: период вне диапазона 1–8; применяются условия без коэффициента нового пользователя.")

    # Скидка по числу лицензий (п.3.4.4): по решению учитывается в блоке «с применением коэффициентов»,
    # а не в базовых условиях.
    k_license_count = 1.0
    license_count_discount_label = ""
    if len(licenses) > 3:
        try:
            lic_disc_df = pd.read_excel(vars_xlsx, sheet_name="Скидки по количеству лицензий")
            k_license_count = discount_by_licenses(lic_disc_df, len(licenses))
        except Exception:
            k_license_count = 1.0
        if k_license_count != 1.0:
            discount_pct = round((1.0 - k_license_count) * 100.0, 2)
            license_count_discount_label = (
                f"{str(k_license_count).replace('.', ',')} (скидка за количество лицензий, {str(discount_pct).replace('.', ',')}%)"
            )
            notes.append(
                f"Скидка по количеству лицензий (п.3.4.4) перенесена в блок расчёта с коэффициентами: {len(licenses)} ВЛ, коэффициент {k_license_count:g}."
            )

    # Условия основного договора:
    # - если применена ветка малого дохода (C3), в базовых условиях показываем стандартную
    #   минималку без C3 (полная базовая), а C3 уходит в ДС;
    # - «гильотина» в базовые условия не включается.
    if small_income_applied:
        standard_min_total, _, _ = compute_min_total(
            licenses=licenses,
            vars_xlsx=vars_xlsx,
            annual_income_for_rules=annual_income_for_rules,
            contract_quarter=args.contract_quarter,
            internet_resources=internet_resources_effective,
            past_year_percent_paid=args.past_year_percent_paid,
            percent_sum_q=None,
            contract_media=args.contract_media,
            use_small_income_branch=False,
            new_user_only=bool(args.new_user),
            assoc_member=bool(args.assoc_member),
            subscriber_total=args.subscriber_total,
            apply_license_count_discount=False,
        )
        base_min_total = round_rub(float(standard_min_total or 0.0))
    else:
        base_min_total = round_rub(float(min_before_guillotine))
    final_percent_sum_q = percent_sum_q

    min_after_alt = round_rub(float(base_min_total) * k_license_count * k42)
    min_with_coeff_y1 = None
    min_with_coeff_y2 = None
    final_min_total = min_after_alt
    if bool(args.new_user):
        min_with_coeff_y1 = round_rub(float(min_after_alt) * 0.75)
        min_with_coeff_y2 = round_rub(float(min_after_alt) * 0.88)
        q_cur = int(args.contract_quarter or 1)
        if 1 <= q_cur <= 4:
            final_min_total = min_with_coeff_y1
        elif 5 <= q_cur <= 8:
            final_min_total = min_with_coeff_y2
        else:
            final_min_total = min_after_alt

    # Условия допсоглашения:
    # - при C3 (малый доход) — специальная минималка на срок до 8 отчётных периодов;
    # - при «гильотине» — минималка по правилам п.3.7.
    addendum_percent_sum_q = None
    addendum_min_total = None
    addendum_min_with_coeff_y1 = None
    addendum_min_with_coeff_y2 = None
    addendum_base_before_coeffs = None
    addendum_base_after_common_coeffs = None
    addendum_limited_8q = False
    addendum_percent_based = False
    applied_coeffs: List[str] = []
    if license_count_discount_label:
        applied_coeffs.append(license_count_discount_label)
    if k42 != 1.0:
        applied_coeffs.append(user_discount_label)
    if bool(args.new_user):
        applied_coeffs.append("0,75 (новый пользователь, 1–4 отчётные периоды)")
        applied_coeffs.append("0,88 (новый пользователь, 5–8 отчётные периоды)")
    if small_income_applied or guillotine_triggered:
        addendum_percent_sum_q = final_percent_sum_q
        if guillotine_triggered:
            addendum_base_before_coeffs = round_rub(float(min_total or 0.0))
            addendum_base_after_alt = round_rub(float(min_total or 0.0) * k_license_count * k42)
            addendum_percent_based = small_income_applied
        else:
            addendum_base_before_coeffs = round_rub(float(min_before_guillotine or 0.0))
            addendum_base_after_alt = round_rub(float(min_before_guillotine or 0.0) * k_license_count * k42)
            addendum_limited_8q = True
        addendum_base_after_common_coeffs = addendum_base_after_alt
        addendum_min_total = addendum_base_after_alt
        if bool(args.new_user):
            addendum_min_with_coeff_y1 = round_rub(float(addendum_base_after_alt) * 0.75)
            addendum_min_with_coeff_y2 = round_rub(float(addendum_base_after_alt) * 0.88)
            q_cur = int(args.contract_quarter or 1)
            if 1 <= q_cur <= 4:
                addendum_min_total = addendum_min_with_coeff_y1
            elif 5 <= q_cur <= 8:
                addendum_min_total = addendum_min_with_coeff_y2
            else:
                addendum_min_total = addendum_base_after_alt

    if final_min_total is not None and final_min_total < 0:
        final_min_total = 0.0

    if final_percent_sum_q is not None and final_percent_sum_q < 0:
        final_percent_sum_q = 0.0

    p.tick("формирую отчёт")
    report = format_report(
        inn=inn,
        year=args.year,
        annual_revenue=args.annual_revenue,
        revenue_q=args.revenue_q,
        expenses_q=args.expenses_q,
        internet_resources=internet_resources_effective,
        contract_quarter=args.contract_quarter,
        new_user_only=bool(args.new_user),
        user_discount_label=user_discount_label,
        fixed_fee_eligible=bool(args.fixed_fee_eligible),
        manual_retransmission=bool(args.manual_retransmission),
        rate_mode=str(args.rate_mode or "license"),
        actual_usage_share=args.actual_usage_share,
        actual_usage_applied=actual_usage_applied,
        actual_usage_category=actual_usage_category,
        licenses=licenses,
        contract_rate=contract_rate,
        base_percent_sum_q=base_percent_sum_q,
        base_min_total=base_min_total,
        final_percent_sum_q=final_percent_sum_q,
        final_min_total=final_min_total,
        min_with_coeff_y1=min_with_coeff_y1,
        min_with_coeff_y2=min_with_coeff_y2,
        addendum_percent_sum_q=addendum_percent_sum_q,
        addendum_min_total=addendum_min_total,
        addendum_min_with_coeff_y1=addendum_min_with_coeff_y1,
        addendum_min_with_coeff_y2=addendum_min_with_coeff_y2,
        addendum_base_before_coeffs=addendum_base_before_coeffs,
        addendum_base_after_common_coeffs=addendum_base_after_common_coeffs,
        k_license_count=k_license_count,
        k42=k42,
        guillotine_triggered=guillotine_triggered,
        addendum_limited_8q=addendum_limited_8q,
        addendum_percent_based=addendum_percent_based,
        applied_coeffs=applied_coeffs,
        volume_coeff_notes=volume_coeff_notes,
        minimum_breakdown=minimum_breakdown,
        notes=notes,
        needs=needs,
        society=(args.society or "РАО"),
    )
    print(report)
    return 0


def run_calc_capture(argv: List[str]) -> Tuple[int, str]:
    """
    Запускает main(argv=...), возвращает (exit_code, stdout_text).
    """
    buf = io.StringIO()
    try:
        with redirect_stdout(buf):
            code = int(main(argv))
    except SystemExit as e:
        code = int(getattr(e, "code", 1) or 0)
    except Exception as e:
        code = 2
        buf.write(f"Ошибка: {type(e).__name__}: {e}\n")
    return code, buf.getvalue()


if __name__ == "__main__":
    raise SystemExit(main())

def fix_mojibake(s: str) -> str:
    """
    Чинит типичный случай: UTF-8 байты были интерпретированы как latin-1.
    Если строка нормальная — вернёт как есть.
    """
    if not s:
        return s
    if "Ð" not in s and "Ñ" not in s:
        return s
    try:
        return s.encode("latin1").decode("utf-8")
    except Exception:
        return s
